"""
Build a Sail Plan workbook from a passage + forecast + buoys + lessons YAML bundle.

Usage:
  python build.py <passage.yaml> <forecast.yaml> <buoys.yaml> <output.xlsx> [lessons.yaml]

If lessons.yaml is omitted, falls back to inputs/lessons.yaml in the same tree.

The workbook is REGENERATED from these inputs every time — there is no
"edit in place" path. To change anything, edit the YAML and rerun.
"""
import sys
import re
import yaml
from pathlib import Path
from openpyxl import Workbook

from sailbuild.compute import build_legs_for_plan
from sailbuild.tabs.plan import render_plan_tab
from sailbuild.tabs.risk_bowtie import render_risk_bowtie
from sailbuild.tabs.briefing import render_briefing
from sailbuild.tabs.support import (
    render_waypoints,
    render_vessel_particulars,
    render_vessel_comparison,
    render_live_buoy_data,
    render_forecast_sources,
    render_buoys_by_wp,
    render_forecast_products,
    render_url_quick_reference,
    render_glossary,
    render_format_reference,
    render_refresh_cadence,
    render_verification_scorecard,
)


class StrictLoader(yaml.SafeLoader):
    """SafeLoader without YAML 1.1's octal-int interpretation.

    Critical for nautical work: '042' is course bearing 42°T, NOT decimal 34
    (octal 042). Default SafeLoader silently does the wrong thing. This loader
    rejects leading-zero octal and loads '042' as the string '042', which
    downstream int coercion handles correctly.
    """
    pass


# Remove the default int resolver and replace with a strict one.
def _install_strict_int_resolver():
    strict_int = re.compile(
        r'^(?:[-+]?(?:0|[1-9][0-9]*)|0x[0-9a-fA-F_]+|0o[0-7_]+|0b[01_]+)$'
    )
    new_resolvers = {}
    for ch, resolvers in yaml.SafeLoader.yaml_implicit_resolvers.items():
        kept = [(tag, regex) for (tag, regex) in resolvers
                if tag != 'tag:yaml.org,2002:int']
        new_resolvers[ch] = kept
    StrictLoader.yaml_implicit_resolvers = new_resolvers
    StrictLoader.add_implicit_resolver(
        'tag:yaml.org,2002:int', strict_int, list('-+0123456789')
    )


_install_strict_int_resolver()


def load_yaml(path):
    with open(path) as f:
        return yaml.load(f, Loader=StrictLoader)


def build(passage_path, forecast_path, buoys_path, output_path, lessons_path=None, status_data=None):
    passage = load_yaml(passage_path)
    forecast = load_yaml(forecast_path)
    buoys = load_yaml(buoys_path)

    if lessons_path is None:
        default_path = Path(passage_path).parent.parent / "lessons.yaml"
        if default_path.exists():
            lessons_path = str(default_path)

    if lessons_path and Path(lessons_path).exists():
        lessons_data = load_yaml(lessons_path)
        lessons = lessons_data.get("lessons", [])
    else:
        lessons = []
        print(f"  ⚠ No lessons.yaml found — Verification Scorecard will be empty.")

    wb = Workbook()
    del wb["Sheet"]

    # Compute legs once; multiple tabs use them
    legs_by_plan = {}
    for plan in passage["plans"]:
        legs_by_plan[plan["id"]] = build_legs_for_plan(passage, forecast, plan["id"])

    # ==========================================================
    # Tab organization — four logical groups:
    #
    #   GROUP 1 — THE ANSWER          : tabs 1 (briefing)
    #   GROUP 2 — THE PLANS           : tabs 2+ (plan + bowtie per scenario)
    #   GROUP 3 — THE WEATHER         : live buoys, forecast sources, etc.
    #   GROUP 4 — ROUTE & VESSEL      : waypoints, vessel particulars
    #   GROUP 5 — STATIC REFERENCE    : format reference, products, glossary,
    #                                    verification scorecard
    #
    # The skipper reads top-to-bottom: "what should I do" → "here's the plan" →
    # "here's the weather that drove it" → "here's the route geography" →
    # "here's the meta info."
    # ==========================================================

    # === GROUP 1: THE ANSWER ===
    # Tab 1: Pre-Departure Briefing — the GO/NO-GO dashboard
    ws = wb.create_sheet("Pre-Departure Briefing")
    render_briefing(ws, passage, forecast, buoys, legs_by_plan)

    # === GROUP 2: THE PLANS ===
    # Tabs 2..2N+1: Plan + Risk Bowtie per departure scenario.
    # This is the actual work product — placed second so the skipper goes
    # briefing → plan immediately without paging past reference material.
    for plan in passage["plans"]:
        sheet_name = plan["tab_label"]
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        ws = wb.create_sheet(sheet_name)
        commentary = forecast.get("plan_commentary", {}).get(plan["id"], "")
        render_plan_tab(
            ws,
            plan_meta=plan,
            commentary=commentary,
            legs=legs_by_plan[plan["id"]],
            total_nm=passage["passage"]["total_nm"],
            passage=passage,
        )

        bowtie_name = f"{plan['tab_label']} Bowtie"
        if len(bowtie_name) > 31:
            bowtie_name = bowtie_name[:31]
        ws = wb.create_sheet(bowtie_name)
        render_risk_bowtie(ws, passage, plan, legs_by_plan[plan["id"]])

    # === GROUP 3: THE WEATHER ===
    # Live buoy obs come first (current ground truth), then forecast attribution
    # per WP, then the buoy-to-WP mapping, then the refresh cadence reminder.
    ws = wb.create_sheet("Live Buoy Data")
    render_live_buoy_data(ws, buoys)

    ws = wb.create_sheet("Forecast Sources by WP")
    render_forecast_sources(ws, passage, forecast)

    ws = wb.create_sheet("Buoys by WP")
    render_buoys_by_wp(ws, passage)

    ws = wb.create_sheet("Refresh Cadence")
    render_refresh_cadence(ws, passage)

    # === GROUP 4: ROUTE & VESSEL ===
    # The "constants" of the passage — geographic route and the boat.
    ws = wb.create_sheet("Waypoints")
    render_waypoints(ws, passage, passage["passage"]["total_nm"])

    ws = wb.create_sheet("Vessel Particulars")
    render_vessel_particulars(ws, passage)

    # OPTIONAL: Vessel Comparison. Only rendered when the passage explicitly
    # opts in via `include_vessel_comparison: true`. Most passages use a
    # single boat; the comparison only adds value when weighing one vessel
    # against another for the same route.
    if passage.get("include_vessel_comparison", False):
        ws = wb.create_sheet("Vessel Comparison")
        render_vessel_comparison(ws, passage, legs_by_plan)

    # === GROUP 5: STATIC REFERENCE ===
    # Methodology and reference material. Placed last because the skipper
    # rarely opens these mid-passage — they're for understanding the
    # methodology or doing post-passage verification.
    ws = wb.create_sheet("Format Reference")
    render_format_reference(ws, passage, forecast, buoys, legs_by_plan, status_data=status_data)

    ws = wb.create_sheet("Forecast Products")
    render_forecast_products(ws, passage)

    ws = wb.create_sheet("URL Quick Reference")
    render_url_quick_reference(ws, passage, buoys)

    ws = wb.create_sheet("Glossary")
    render_glossary(ws, passage)

    ws = wb.create_sheet("Verification Scorecard")
    render_verification_scorecard(ws, passage, lessons)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return legs_by_plan


if __name__ == "__main__":
    import argparse
    import json
    ap = argparse.ArgumentParser(description="Build a Sail Plan workbook")
    ap.add_argument("passage", help="passage YAML")
    ap.add_argument("forecast", help="forecast YAML")
    ap.add_argument("buoys", help="buoys YAML")
    ap.add_argument("output", help="output xlsx path")
    ap.add_argument("lessons", nargs="?", help="optional lessons YAML")
    ap.add_argument("--require", metavar="STATUS_JSON",
                    help="Refuse to build if any required source in this "
                         "weather_pull _status.json is not covered (FRESH/STALE).")
    ap.add_argument("--allow-stale", action="store_true",
                    help="With --require, treat ARCHIVED/missing as warning, not block.")
    args = ap.parse_args()

    # Freshness gate
    status = None
    if args.require:
        status_path = Path(args.require)
        if not status_path.exists():
            print(f"✗ --require status file not found: {status_path}")
            print(f"  Run weather_pull.py first to generate it.")
            sys.exit(2)
        status = json.loads(status_path.read_text())
        missing = [s for s in status["sources"]
                   if s.get("required") and not s.get("covered")]
        if missing:
            print(f"\n✗ BUILD BLOCKED — {len(missing)} required source(s) not covered:")
            for s in missing:
                print(f"    {s['id']} ({s['kind']}) — last attempts: ", end="")
                tiers = [a.get("tier") or "fail" for a in s.get("attempts", [])]
                print(", ".join(tiers) if tiers else "(none)")
            if args.allow_stale:
                print("  (--allow-stale set: continuing anyway)\n")
            else:
                print("\n  Pass --allow-stale to build on incomplete data, or "
                      "ingest pastes and rerun.")
                sys.exit(2)
        else:
            n = len([s for s in status["sources"] if s.get("covered")])
            print(f"✓ Freshness gate passed: {n} sources covered (status: {status_path.name})")

    legs = build(args.passage, args.forecast, args.buoys, args.output, args.lessons,
                 status_data=status)
    print(f"\n✓ Built workbook: {args.output}")
    for plan_id, plan_legs in legs.items():
        arrival = plan_legs[-1]
        print(f"  {plan_id}: arrive {arrival.eta_str}, {arrival.cum_time_hr:.1f} hr "
              f"({arrival.cum_sailing_hr:.1f} sail / {arrival.cum_motoring_hr:.1f} motor), color {arrival.eta_color}")

    # === Route files (KML for Google Earth/My Maps/OpenCPN, GPX for plotters) ===
    import yaml as _yaml
    from sailbuild.export import write_route_files
    with open(args.passage) as _f:
        _passage = _yaml.safe_load(_f)
    _output_dir = str(Path(args.output).parent)
    kml_path, gpx_path = write_route_files(_passage, _output_dir)
    print(f"✓ Route files:")
    print(f"  KML: {kml_path}")
    print(f"  GPX: {gpx_path}")

    # === STANDING PROCEDURE: End-of-run verification pass ===
    # Reads back the produced workbook and runs cell-level consistency checks
    # (vessel designation leaks, polar grid mismatch, arrival timing colors,
    # buoy coordinate sanity, tab inventory). Anything found here means the
    # workbook should NOT be delivered until fixed.
    from sailbuild.verify import print_verification_report
    with open(args.forecast) as _f:
        _forecast = _yaml.safe_load(_f)
    with open(args.buoys) as _f:
        _buoys = _yaml.safe_load(_f)
    _errs = print_verification_report(args.output, _passage, _forecast, _buoys)
    if _errs > 0:
        print(f"\n  ⚠ {_errs} error(s) found in verification — review before delivering workbook.")
