"""
Shared styles for sailbuild workbooks.

Centralized so every tab uses the same typography, colors, borders, and
alignment. The goal is presentation-level polish: numbers right-aligned with
appropriate precision, text left-aligned and wrapped, codes/degrees centered,
consistent header style throughout.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ======================================================================
# COLOR PALETTE — presentation-grade, restrained
# ======================================================================
COLOR_HEADER_FILL = "305496"       # Dark blue (table column headers)
COLOR_HEADER_FONT = "FFFFFF"       # White on dark blue
COLOR_SUBHEADER_FILL = "D9E1F2"    # Light blue (section bands)
COLOR_TITLE_FONT = "1F3864"        # Deep navy for page titles

COLOR_GOOD_FILL = "C6EFCE"         # Excel "Good" green
COLOR_NEUTRAL_FILL = "FFEB9C"      # Excel "Neutral" yellow
COLOR_BAD_FILL = "FFC7CE"          # Excel "Bad" red
COLOR_DANGER_FONT = "9C0006"       # Dark red
COLOR_GOOD_FONT = "006100"         # Dark green
COLOR_NEUTRAL_FONT = "9C5700"      # Dark amber

COLOR_BORDER = "BFBFBF"            # Subtle gray
COLOR_BORDER_HEAVY = "808080"      # Section separator
COLOR_BAND_LIGHT = "F2F2F2"        # Alternating-row band
COLOR_NOTE_FONT = "595959"         # Italic notes / captions


# ======================================================================
# FONT FAMILY — Calibri throughout for screen + print parity
# ======================================================================
FONT_NAME = "Calibri"


# ======================================================================
# FONTS — by role
# ======================================================================
def page_title_font():
    """Top of each tab — large, navy, bold."""
    return Font(name=FONT_NAME, size=16, bold=True, color=COLOR_TITLE_FONT)


def page_subtitle_font():
    """Caption/description directly under page title."""
    return Font(name=FONT_NAME, size=10, italic=True, color=COLOR_NOTE_FONT)


def section_header_font():
    """In-tab section heading — bold dark navy on light fill."""
    return Font(name=FONT_NAME, size=12, bold=True, color=COLOR_TITLE_FONT)


def table_header_font():
    """Column header — white on dark blue."""
    return Font(name=FONT_NAME, size=10, bold=True, color=COLOR_HEADER_FONT)


def body_font():
    """Standard body text."""
    return Font(name=FONT_NAME, size=11)


def body_bold_font():
    return Font(name=FONT_NAME, size=11, bold=True)


def caption_font():
    """Small italic for footnotes / legends."""
    return Font(name=FONT_NAME, size=9, italic=True, color=COLOR_NOTE_FONT)


# ======================================================================
# FILLS
# ======================================================================
def fill(hex_color):
    """Solid fill in a hex color (no '#' prefix)."""
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def header_fill():
    return fill(COLOR_HEADER_FILL)


def subheader_fill():
    return fill(COLOR_SUBHEADER_FILL)


def band_fill():
    return fill(COLOR_BAND_LIGHT)


# ======================================================================
# BORDERS
# ======================================================================
def thin_border():
    s = Side(style="thin", color=COLOR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def heavy_bottom_border():
    """For separator rows beneath section headers."""
    return Border(
        left=Side(style="thin", color=COLOR_BORDER),
        right=Side(style="thin", color=COLOR_BORDER),
        top=Side(style="thin", color=COLOR_BORDER),
        bottom=Side(style="medium", color=COLOR_BORDER_HEAVY),
    )


def no_border():
    return Border()


# ======================================================================
# ALIGNMENT — by data type
# ======================================================================
def align_center():
    """Codes, degrees, short labels."""
    return Alignment(horizontal="center", vertical="center", wrap_text=False)


def align_number():
    """Numeric data, right-aligned, vertically centered."""
    return Alignment(horizontal="right", vertical="center", wrap_text=False)


def align_text_left():
    """Long text, left-aligned, wraps."""
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def align_text_top():
    """Multi-line text in a tall cell — anchors to top."""
    return Alignment(horizontal="left", vertical="top", wrap_text=True)


def align_header():
    """Table column headers — centered, wrapped."""
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


# ======================================================================
# COMPOSITE STYLERS — apply complete formatting in one call
# ======================================================================
def style_page_title(cell, text):
    """Apply page-title style to a cell + write text."""
    cell.value = text
    cell.font = page_title_font()
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_page_subtitle(cell, text):
    cell.value = text
    cell.font = page_subtitle_font()
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_section_header(cell, text):
    """In-tab section heading — bold, navy text, subheader fill."""
    cell.value = text
    cell.font = section_header_font()
    cell.fill = subheader_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = heavy_bottom_border()


def style_table_header(cell, text=None):
    """Column header style: white on dark blue, centered, wrapped, bordered."""
    if text is not None:
        cell.value = text
    cell.font = table_header_font()
    cell.fill = header_fill()
    cell.alignment = align_header()
    cell.border = thin_border()


# Back-compat alias for older code
def style_header_cell(cell):
    style_table_header(cell)


def style_number_cell(cell, value=None, number_format="0.0", fill_color=None):
    """Right-aligned numeric cell. number_format examples: '0', '0.0', '0.00', '0.0\"%\"'."""
    if value is not None:
        cell.value = value
    cell.font = body_font()
    cell.alignment = align_number()
    cell.number_format = number_format
    cell.border = thin_border()
    if fill_color:
        cell.fill = fill(fill_color)


def style_centered_cell(cell, value=None, fill_color=None, bold=False):
    """Center-aligned cell — for course/TWA/angle codes."""
    if value is not None:
        cell.value = value
    cell.font = body_bold_font() if bold else body_font()
    cell.alignment = align_center()
    cell.border = thin_border()
    if fill_color:
        cell.fill = fill(fill_color)


def style_text_cell(cell, value=None, fill_color=None, wrap=True):
    """Left-aligned text cell. Use for descriptions, names, free-form notes."""
    if value is not None:
        cell.value = value
    cell.font = body_font()
    cell.alignment = align_text_left() if wrap else Alignment(horizontal="left", vertical="center")
    cell.border = thin_border()
    if fill_color:
        cell.fill = fill(fill_color)


def style_data_cell(cell, fill_color=None):
    """Back-compat alias — centered cell with optional fill."""
    cell.alignment = align_center()
    cell.border = thin_border()
    if fill_color:
        cell.fill = fill(fill_color)


# ======================================================================
# WHOLE-ROW BANDING — for visual rhythm in dense tables
# ======================================================================
def apply_row_band(ws, row, start_col, end_col, fill_color=None):
    """Apply a banded fill across cells start_col..end_col in `row`.

    Used for the row_band coloring on Plan tabs (whole-row risk color) and
    for alternating row stripes in dense reference tables.
    """
    if fill_color is None:
        return
    f = fill(fill_color)
    for c in range(start_col, end_col + 1):
        ws.cell(row, c).fill = f


def freeze_at(ws, cell_ref):
    """Freeze panes at a given anchor cell (everything above/left stays visible)."""
    ws.freeze_panes = cell_ref


def set_column_widths(ws, widths_dict):
    """Bulk-set column widths. widths_dict = {'A': 12, 'B': 24, ...}."""
    for col, w in widths_dict.items():
        ws.column_dimensions[col].width = w


# ======================================================================
# Legacy aliases (back-compat with old callers)
# ======================================================================
def header_font():
    return table_header_font()


def header_alignment():
    return align_header()


def cell_alignment():
    return align_center()


def wrap_alignment():
    return align_text_top()
