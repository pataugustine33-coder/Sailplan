"""
Pre-Departure Briefing tab — single-screen synthesis for GO/NO-GO decision.

Purpose: when you're 4-6 hours pre-departure and need to make a decision, this
is the ONE tab you look at. Synthesizes:
  - Departure scenarios side-by-side with arrival color
  - Plan character classification (sailing-heavy vs motor-heavy)
  - Critical operational gates (SCAs in effect, frontal passage timing, etc.)
  - Forecast cycle freshness summary
  - Buoy ground truth snapshot
  - Outstanding decision triggers

Designed to be readable at-a-glance on iPad/phone screens.
"""
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from ..styles import (
    style_header_cell, fill,
    COLOR_GOOD_FILL, COLOR_NEUTRAL_FILL, COLOR_BAD_FILL,
    COLOR_SUBHEADER_FILL, COLOR_HEADER_FILL, COLOR_HEADER_FONT,
    COLOR_DANGER_FONT, COLOR_GOOD_FONT,
)


def render_briefing(ws, passage, forecast, buoys, legs_by_plan):
    """Single-screen pre-departure briefing."""
    # Wide first column for labels; plan columns sized for body text
    ws.column_dimensions["A"].width = 26
    plan_count = len(passage["plans"])
    for i, plan in enumerate(passage["plans"], 2):
        col_letter = ws.cell(1, i).column_letter
        ws.column_dimensions[col_letter].width = 44

    row = 1

    # === BLOCK 1: Title + cycle freshness ===
    title_cell = ws.cell(row, 1, value=f"PRE-DEPARTURE BRIEFING — {passage['passage']['name']}")
    title_cell.font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    title_cell.fill = fill(COLOR_HEADER_FILL)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1 + plan_count)
    ws.row_dimensions[row].height = 38
    row += 1

    cycle = forecast["cycle"]
    cwf_parts = [f"CWF{office} {cd['label_short']}" for office, cd in cycle.items()
                 if office.startswith("cwf") and isinstance(cd, dict)]
    afd_parts = [f"AFD{office[3:].upper()} {cd['label_short']}"
                 for office, cd in cycle.items()
                 if office.startswith("afd") and isinstance(cd, dict)]
    cwf_label = " | ".join(f"CWF{o.upper()} {cd['label_short']}"
                            for o, cd in [(k[3:], v) for k, v in cycle.items()
                                          if k.startswith("cwf") and isinstance(v, dict)])
    afd_label = " | ".join(f"AFD{o.upper()} {cd['label_short']}"
                            for o, cd in [(k[3:], v) for k, v in cycle.items()
                                          if k.startswith("afd") and isinstance(v, dict)])
    cycle_text = f"Cycle: {cwf_label}    AFDs: {afd_label}    Buoys: {buoys['pull']['label']}"
    c = ws.cell(row, 1, value=cycle_text)
    c.font = Font(size=9, italic=True, color="595959")
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1 + plan_count)
    ws.row_dimensions[row].height = 28
    row += 2  # blank row gap

    # === BLOCK 2: Plan headers (color-coded by arrival) ===
    ws.cell(row, 1, value="").fill = fill(COLOR_SUBHEADER_FILL)
    for i, plan in enumerate(passage["plans"], 2):
        legs = legs_by_plan[plan["id"]]
        arrival_color = legs[-1].eta_color if legs else "FFFFFF"
        c = ws.cell(row, i, value=plan["tab_label"])
        c.font = Font(size=12, bold=True)
        c.fill = fill(arrival_color)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 26
    row += 1

    # === BLOCK 3: Departure / Arrival / Duration ===
    blocks = [
        ("Departure", lambda l, p: f"{p['depart_day']} {_h(p['depart_hour'])} EDT"),
        ("Arrival ETA", lambda l, p: l[-1].eta_str),
        ("Arrival lighting", lambda l, p: _arrival_lighting(l[-1].eta_color)),
        ("Passage duration", lambda l, p: f"{l[-1].cum_time_hr:.1f} hr"),
        ("Average SOG", lambda l, p: f"{passage['passage']['total_nm']/l[-1].cum_time_hr:.1f} kt"),
        ("Sailing time", lambda l, p: f"{l[-1].cum_sailing_hr:.1f} hr ({100*l[-1].cum_sailing_hr/l[-1].cum_time_hr:.0f}%)"),
        ("Motoring time", lambda l, p: f"{l[-1].cum_motoring_hr:.1f} hr ({100*l[-1].cum_motoring_hr/l[-1].cum_time_hr:.0f}%)"),
        ("Plan character", lambda l, p: _plan_character(l)),
    ]
    for label, fn in blocks:
        ws.cell(row, 1, value=label).font = Font(bold=True)
        ws.cell(row, 1).fill = fill(COLOR_SUBHEADER_FILL)
        ws.cell(row, 1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for i, plan in enumerate(passage["plans"], 2):
            legs = legs_by_plan[plan["id"]]
            c = ws.cell(row, i, value=fn(legs, plan))
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1
    row += 1  # gap

    # === BLOCK 4: Critical operational gates ===
    ws.cell(row, 1, value="OPERATIONAL GATES").font = Font(size=12, bold=True, color="FFFFFF")
    ws.cell(row, 1).fill = fill(COLOR_HEADER_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1 + plan_count)
    ws.row_dimensions[row].height = 22
    row += 1

    gates = _build_gates(passage, forecast, buoys, legs_by_plan)
    for gate_label, gate_per_plan in gates:
        ws.cell(row, 1, value=gate_label).font = Font(bold=True)
        ws.cell(row, 1).alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)
        for i, plan in enumerate(passage["plans"], 2):
            verdict, detail = gate_per_plan[plan["id"]]
            c = ws.cell(row, i, value=detail)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if verdict == "ok":
                c.fill = fill(COLOR_GOOD_FILL)
            elif verdict == "watch":
                c.fill = fill(COLOR_NEUTRAL_FILL)
            elif verdict == "stop":
                c.fill = fill(COLOR_BAD_FILL)
        ws.row_dimensions[row].height = 48
        row += 1
    row += 1

    # === BLOCK 5: Buoy ground truth ===
    ws.cell(row, 1, value="BUOY GROUND TRUTH").font = Font(size=12, bold=True, color="FFFFFF")
    ws.cell(row, 1).fill = fill(COLOR_HEADER_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1 + plan_count)
    ws.row_dimensions[row].height = 22
    row += 1

    for st in buoys["stations"]:
        if st.get("status") == "offline":
            continue
        label = f"Buoy {st['id']} — {st['name']}"
        ws.cell(row, 1, value=label).font = Font(bold=True)
        ws.cell(row, 1).alignment = Alignment(horizontal="left", indent=1, wrap_text=True)
        if "wave_summary" in st:
            obs = st["wave_summary"]
        else:
            wind = st.get("wind_dir", "?")
            kt = st.get("wind_kt", "?")
            gust = st.get("gust_kt", "?")
            press = st.get("pressure_inhg", "?")
            obs = f"{wind} {kt} kt, gust {gust} kt, {press} inHg ({st.get('reading_time', '?')})"
        c = ws.cell(row, 2, value=obs)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=1 + plan_count)
        ws.row_dimensions[row].height = 28
        row += 1
    row += 1

    # === BLOCK 6: Critical synthesis findings ===
    ws.cell(row, 1, value="KEY FINDINGS THIS CYCLE").font = Font(size=12, bold=True, color="FFFFFF")
    ws.cell(row, 1).fill = fill(COLOR_HEADER_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1 + plan_count)
    ws.row_dimensions[row].height = 22
    row += 1

    findings = buoys.get("findings", "").strip()
    if findings:
        ws.cell(row, 1, value="From buoy verification:").font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1 + plan_count)
        row += 1
        c = ws.cell(row, 1, value=findings)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1 + plan_count)
        ws.row_dimensions[row].height = 140
        row += 2

    # AFD-driven findings (key messages from each office)
    ws.cell(row, 1, value="AFD marine guidance (current cycle):").font = Font(bold=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1 + plan_count)
    row += 1
    afd_text = "\n\n".join(
        f"{office}: {text.strip()}"
        for office, text in forecast.get("afd_marine", {}).items()
    )
    c = ws.cell(row, 1, value=afd_text)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1 + plan_count)
    ws.row_dimensions[row].height = 280
    row += 2

    # === BLOCK 7: Decision triggers (what would change my mind?) ===
    ws.cell(row, 1, value="DECISION TRIGGERS (what would change my mind)").font = Font(size=12, bold=True, color="FFFFFF")
    ws.cell(row, 1).fill = fill(COLOR_HEADER_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1 + plan_count)
    ws.row_dimensions[row].height = 22
    row += 1

    triggers = _build_triggers(passage, forecast, buoys, legs_by_plan)
    for trigger_text in triggers:
        c = ws.cell(row, 1, value=trigger_text)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1 + plan_count)
        ws.row_dimensions[row].height = 38
        row += 1


def _h(decimal_h):
    """Format decimal hour as 'H:MM AM/PM'."""
    h = int(decimal_h)
    m = int(round((decimal_h - h) * 60))
    if m == 60:
        h += 1
        m = 0
    if h == 0:
        return f"12:{m:02d} AM"
    if h < 12:
        return f"{h}:{m:02d} AM"
    if h == 12:
        return f"12:{m:02d} PM"
    return f"{h-12}:{m:02d} PM"


def _arrival_lighting(color):
    return {
        "C6EFCE": "DAY ✓",
        "FFEB9C": "TWILIGHT / DAWN",
        "FFC7CE": "⚠ NIGHT",
    }.get(color, "")


def _plan_character(legs):
    """Classify plan as sailing-heavy / mixed / motor-heavy based on sail %."""
    arr = legs[-1]
    total = arr.cum_time_hr
    sail = arr.cum_sailing_hr
    if total <= 0:
        return "?"
    pct = 100 * sail / total
    if pct >= 70:
        return f"SAILING-HEAVY ({pct:.0f}% sail)"
    elif pct >= 30:
        return f"MIXED ({pct:.0f}% sail)"
    else:
        return f"MOTOR-HEAVY ({pct:.0f}% sail)"


def _build_gates(passage, forecast, buoys, legs_by_plan):
    """Build operational gate evaluations per plan.

    Each gate returns (verdict, detail) where verdict is "ok", "watch", or "stop".
    """
    gates = []

    # Gate 1: SCA status at departure
    def sca_gate():
        result = {}
        for plan in passage["plans"]:
            legs = legs_by_plan[plan["id"]]
            wp0 = legs[0] if legs else None
            if not wp0:
                result[plan["id"]] = ("watch", "no data")
                continue
            # Check the zone the WP0 sits in for SCA at departure time
            assignment = forecast["waypoint_assignments"][plan["id"]].get("WP0", {})
            zone = assignment.get("zone")
            sca_info = forecast["zones"].get(zone, {}).get("SCA")
            if sca_info:
                result[plan["id"]] = (
                    "watch",
                    f"SCA in {zone} until {sca_info['in_effect_until']}. {sca_info.get('reason', '')}"
                )
            else:
                result[plan["id"]] = ("ok", "No SCA active at departure zone.")
        return result

    # Gate 2: Arrival timing window
    def arrival_gate():
        result = {}
        at = passage["arrival_timing"]
        pref_start, pref_end = at["preferred_window"]["start"], at["preferred_window"]["end"]
        for plan in passage["plans"]:
            legs = legs_by_plan[plan["id"]]
            arr = legs[-1] if legs else None
            if not arr:
                result[plan["id"]] = ("watch", "no data")
                continue
            h = arr.eta_decimal_hour
            if pref_start <= h <= pref_end:
                result[plan["id"]] = ("ok", f"{arr.eta_str} — within preferred window {_h(pref_start)}-{_h(pref_end)}")
            elif arr.eta_color == "C6EFCE":
                result[plan["id"]] = ("ok", f"{arr.eta_str} — DAY (outside preferred but safe daylight)")
            elif arr.eta_color == "FFEB9C":
                result[plan["id"]] = ("watch", f"{arr.eta_str} — TWILIGHT (light increasing rapidly; monitor)")
            else:
                result[plan["id"]] = ("stop", f"{arr.eta_str} — NIGHT (avoid; consider departure shift)")
        return result

    # Gate 3: Peak wind / sea exposure
    def exposure_gate():
        result = {}
        # Use the actual vessel designation from passage YAML (not hard-coded HR 48).
        vessel_label = passage.get("vessel", {}).get("designation", "this vessel")
        for plan in passage["plans"]:
            legs = legs_by_plan[plan["id"]]
            sail_legs = [l for l in legs if l.course_out is not None and not l.is_motor]
            all_legs = [l for l in legs if l.course_out is not None]
            if not all_legs:
                result[plan["id"]] = ("watch", "no data")
                continue
            peak_wind = max(l.wind_kt_high for l in all_legs)
            peak_sea = max(l.sea_ft_high for l in all_legs)
            if peak_wind >= 25 or peak_sea >= 8:
                result[plan["id"]] = ("stop", f"Peak {peak_wind:.0f} kt / {peak_sea:.0f} ft — heavy weather; reconsider")
            elif peak_wind >= 18 or peak_sea >= 6:
                result[plan["id"]] = ("watch", f"Peak {peak_wind:.0f} kt / {peak_sea:.0f} ft — sporty; manageable for {vessel_label}")
            else:
                result[plan["id"]] = ("ok", f"Peak {peak_wind:.0f} kt / {peak_sea:.0f} ft — moderate")
        return result

    # Gate 4: Frontal passage during transit (signal = pressure trend, not wind text)
    def frontal_gate():
        result = {}
        for plan in passage["plans"]:
            legs = legs_by_plan[plan["id"]]
            # Frontal signal: pressure trend "Falling fast" or "Bottoming" at any WP
            frontal_wps = []
            for leg in legs:
                if leg.pressure_trend and any(
                    flag in leg.pressure_trend
                    for flag in ("Falling fast", "Bottoming")
                ):
                    frontal_wps.append(leg)
            if frontal_wps:
                first = frontal_wps[0]
                result[plan["id"]] = (
                    "watch",
                    f"Frontal passage at {first.wp_id} ({first.eta_str}). "
                    f"Pressure {first.pressure_trend.lower()}. Verify front timing at "
                    f"nearest buoy and REEF BEFORE the front."
                )
            else:
                result[plan["id"]] = ("ok", "No frontal passage during transit")
        return result

    # Gate 5: Single-handed sustainability (long motoring + long night)
    def single_hander_gate():
        result = {}
        for plan in passage["plans"]:
            legs = legs_by_plan[plan["id"]]
            arr = legs[-1] if legs else None
            if not arr:
                result[plan["id"]] = ("watch", "no data")
                continue
            total = arr.cum_time_hr
            if total > 30:
                result[plan["id"]] = ("watch", f"{total:.0f} hr passage — single-handed fatigue territory")
            elif total > 24:
                result[plan["id"]] = ("watch", f"{total:.0f} hr — manageable single-handed with watch discipline")
            else:
                result[plan["id"]] = ("ok", f"{total:.0f} hr — single-handed comfortable range")
        return result

    gates.append(("SCA / advisory status", sca_gate()))
    gates.append(("Arrival timing window", arrival_gate()))
    gates.append(("Peak wind / sea exposure", exposure_gate()))
    gates.append(("Frontal passage during transit", frontal_gate()))
    gates.append(("Single-handed sustainability", single_hander_gate()))

    return gates


def _build_triggers(passage, forecast, buoys, legs_by_plan):
    """Build the 'what would change my mind' bullet list."""
    triggers = []

    # Generic decision triggers — could be parameterized in passage.yaml later
    triggers.append(
        "⚠️ If next buoy pull shows wind 5+ kt above forecast at departure zone → expect rougher passage; "
        "validate against AFD narrative before departure."
    )
    triggers.append(
        "⚠️ If AFD updates with SCA extension into departure window → DELAY departure until SCA expires + 2 hr buffer."
    )
    triggers.append(
        "⚠️ If pressure trace at departure-zone buoy is still FALLING (not bottoming/rising) → front not fully cleared; "
        "delay departure 2-4 hr."
    )
    triggers.append(
        "⚠️ If forecast revises arrival ETA into NIGHT (red) window → use Reverse Calculator on Format Reference tab "
        "to compute required departure shift."
    )
    triggers.append(
        "✅ Re-pull cycle at: next 4-AM cycle (final pre-departure check), then again at departure -2 hr for "
        "current-state buoy snapshot."
    )

    return triggers
