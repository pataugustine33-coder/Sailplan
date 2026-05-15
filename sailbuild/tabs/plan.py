"""
Plan tab renderer — the primary deliverable.

Layout (columns A-AB = 28 columns):
  A  WP                 (5)    centered
  B  Description        (24)   left-aligned, wraps
  C  Cum NM             (8)    right-aligned, 0.0
  D  ETA (EDT)          (16)   centered, color-coded by daylight window
  E  Pressure (inHg)    (10)   right-aligned, 0.00
  F  Pressure Trend     (14)   centered, color-coded
  G  Wind Dir           (16)   centered (e.g. "050° NE")
  H  Wind kt            (10)   centered (e.g. "5–10")
  I  Gust kt            (8)    centered (e.g. "g25")
  J  Sea Ht (ft)        (10)   centered
  K  Period (sec)       (9)    centered
  L  Course (°T)        (10)   centered
  M  TWA (°)            (8)    centered
  N  Sea From (°)       (16)   centered, color-coded by source
  O  Sea Angle / Pos    (14)   centered
  P  Polar Speed (kt)   (11)   right-aligned, 0.0
  Q  Boat Speed (kt)    (11)   right-aligned, 0.0
  R  AWS (kt)           (9)    right-aligned, 0.0
  S  AWA (°)            (8)    centered
  T  Sail Mode          (28)   left-aligned, wraps
  U  Wind/Sea Rose      (18)   embedded PNG (120×120) — boat geometry
  V  Polar @ TWS        (18)   embedded PNG (120×120) — boat polar at this TWS with TWA marker
  W  Notes              (60)   left-aligned, wraps, top-anchored
  X  Weather Risk       (40)   left-aligned, wraps, color-coded fill
  Y  Sea Source         (38)   left-aligned, wraps, italic
  Z  Cum Time (hr)      (11)   right-aligned, 0.0
  AA Cum Sailing (hr)   (11)   right-aligned, 0.0
  AB Cum Motoring (hr)  (11)   right-aligned, 0.0

Frozen panes at D4 — WP/Description/Cum NM/ETA stay visible when scrolling
right; column headers stay visible when scrolling down.
"""
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as XLImage
from ..styles import (
    style_page_title, style_page_subtitle, style_table_header,
    style_number_cell, style_centered_cell, style_text_cell,
    fill, body_bold_font, caption_font, thin_border, freeze_at, set_column_widths,
    align_center, align_number, align_text_top,
    COLOR_GOOD_FILL, COLOR_NEUTRAL_FILL, COLOR_BAD_FILL,
    COLOR_SUBHEADER_FILL, COLOR_TITLE_FONT, COLOR_NOTE_FONT,
)
from ..compute import ROW_BAND_COLORS
from ..rose import rose_png_bytes, HAVE_CAIROSVG


# Column definitions: (header_text, width, align_type)
# align_type: "center" | "number" | "text" — drives alignment + number format
COLUMNS = [
    ("WP",                 5,  "center"),     # A
    ("Description",        24, "text"),       # B
    ("Cum NM",             8,  "number"),     # C
    ("ETA (EDT)",          16, "center"),     # D
    ("Pressure (inHg)",    10, "number"),     # E
    ("Pressure Trend",     14, "center"),     # F
    ("Wind Dir",           16, "center"),     # G
    ("Wind kt",            10, "center"),     # H
    ("Gust kt",            8,  "center"),     # I
    ("Sea Ht (ft)",        10, "center"),     # J
    ("Period (sec)",       9,  "center"),     # K
    ("Course (°T)",        10, "center"),     # L
    ("TWA (°)",            8,  "center"),     # M
    ("Sea From (°)",       16, "center"),     # N
    ("Sea Angle / Pos",    14, "center"),     # O
    ("Polar Speed (kt)",   11, "number"),     # P
    ("Boat Speed (kt)",    11, "number"),     # Q
    ("AWS (kt)",           9,  "number"),     # R
    ("AWA (°)",            8,  "center"),     # S
    ("Sail Mode",          28, "text"),       # T
    ("Wind/Sea Rose",      18, "center"),     # U
    ("Polar @ TWS",        18, "center"),     # V — boat polar curve at this TWS, TWA marker
    ("Notes",              60, "text"),       # W
    ("Weather Risk",       40, "text"),       # X
    ("Sea Source",         38, "text"),       # Y
    ("Cum Time (hr)",      11, "number"),     # Z
    ("Cum Sailing (hr)",   11, "number"),     # AA
    ("Cum Motoring (hr)",  11, "number"),     # AB
]
N_COLS = len(COLUMNS)  # 28


def render_plan_tab(ws, plan_meta: dict, commentary: str, legs: list,
                    total_nm: float, passage: dict = None):
    """Write a complete Plan tab into worksheet `ws`."""

    # === Row 1: Title ===
    title = f"{plan_meta['tab_label']} {plan_meta.get('qualifier', '')}".strip()
    style_page_title(ws.cell(1, 1), title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N_COLS)
    ws.row_dimensions[1].height = 26

    # === Row 2: Description / commentary ===
    desc = " ".join(commentary.split())
    style_page_subtitle(ws.cell(2, 1), desc)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N_COLS)
    # Approximate height by content length (4 chars per row at width ~600 chars wide)
    ws.row_dimensions[2].height = max(60, min(140, len(desc) // 6))

    # === Row 3: Column headers ===
    col_widths = {}
    for col_idx, (header, width, _kind) in enumerate(COLUMNS, 1):
        cell = ws.cell(3, col_idx)
        style_table_header(cell, header)
        # Map col_idx to letter for width-setting
        col_letter = ws.cell(3, col_idx).column_letter
        col_widths[col_letter] = width
    set_column_widths(ws, col_widths)
    ws.row_dimensions[3].height = 36

    # === Data rows (row 4+) ===
    sailing_total = 0.0
    motoring_total = 0.0
    total_time = 0.0

    for leg_idx, leg in enumerate(legs):
        r = 4 + leg_idx
        _write_leg_row(ws, r, leg)
        _embed_rose(ws, r, leg)
        _embed_mini_polar(ws, r, leg, passage)

        if leg.course_out is not None:
            if leg.is_motor:
                motoring_total += leg.leg_hours
            else:
                sailing_total += leg.leg_hours
            total_time += leg.leg_hours

    # === Totals row ===
    totals_row = 4 + len(legs)
    _write_totals_row(ws, totals_row, total_nm, total_time, sailing_total, motoring_total)

    # === Wind/sea timeline strip ===
    # Visual summary of conditions across the passage — bars for wind kt,
    # line for sea height, gust marks where present. Sits below the totals
    # row as the "at-a-glance" picture of what the passage will feel like.
    timeline_anchor_row = totals_row + 2
    _embed_timeline_strip(ws, timeline_anchor_row, legs)

    # === Legend block ===
    legend_anchor = timeline_anchor_row + 19  # leave room for the taller chart (340 px ≈ 17 rows)
    last_legend_row = _write_legend_block(ws, legend_anchor)

    # === Footer note (vessel polar reference) ===
    _write_footer_note(ws, last_legend_row + 2, passage)

    # === Freeze panes — column headers + first 3 ID columns stay visible ===
    # D4 means: rows 1-3 (titles + headers) frozen at top,
    #          columns A-C (WP, Description, Cum NM) frozen at left
    freeze_at(ws, "D4")


# ======================================================================
# Row writers
# ======================================================================
def _write_leg_row(ws, r, leg):
    """Write all 27 columns for a single Leg."""

    # A: WP id, centered, bold
    style_centered_cell(ws.cell(r, 1), leg.wp_id, bold=True)

    # B: Description (WP name), left-aligned with wrap
    style_text_cell(ws.cell(r, 2), leg.wp_name)

    # C: Cum NM, right-aligned 0.0
    style_number_cell(ws.cell(r, 3), leg.cum_nm, number_format="0.0")

    # D: ETA — center-aligned, ETA color fill
    style_centered_cell(ws.cell(r, 4), leg.eta_str, fill_color=leg.eta_color, bold=True)

    # E: Pressure (inHg), 0.00
    if leg.pressure_inhg is not None:
        style_number_cell(ws.cell(r, 5), leg.pressure_inhg, number_format="0.00")
    else:
        style_centered_cell(ws.cell(r, 5), "")

    # F: Pressure Trend — color-coded
    if leg.pressure_trend:
        trend_lower = leg.pressure_trend.lower()
        if "falling fast" in trend_lower or "bottoming" in trend_lower:
            fc = COLOR_BAD_FILL
        elif trend_lower.startswith("falling"):
            fc = COLOR_NEUTRAL_FILL
        elif "rising" in trend_lower or "steady" in trend_lower:
            fc = COLOR_GOOD_FILL
        else:
            fc = None
        style_centered_cell(ws.cell(r, 6), leg.pressure_trend, fill_color=fc)
    else:
        style_centered_cell(ws.cell(r, 6), "")

    # G-K: Wind block (rendered same for mid-route and terminal WPs)
    if leg.wind_dir_deg and leg.wind_kt_low:
        wind_dir_str = f"{leg.wind_dir_deg:03d}° {leg.wind_dir_text}"
    else:
        wind_dir_str = leg.wind_dir_text or "—"
    style_centered_cell(ws.cell(r, 7), wind_dir_str)

    style_centered_cell(
        ws.cell(r, 8),
        f"{leg.wind_kt_low:g}–{leg.wind_kt_high:g}" if leg.wind_kt_low else "—",
    )
    style_centered_cell(
        ws.cell(r, 9),
        f"g{leg.gust_kt:g}" if leg.gust_kt else "—",
    )
    style_centered_cell(
        ws.cell(r, 10),
        f"{leg.sea_ft_low:g}–{leg.sea_ft_high:g}" if leg.sea_ft_low else "—",
    )
    if leg.sea_period_s:
        per = int(leg.sea_period_s) if leg.sea_period_s == int(leg.sea_period_s) else leg.sea_period_s
        style_centered_cell(ws.cell(r, 11), per)
    else:
        style_centered_cell(ws.cell(r, 11), "—")

    # L-S: Per-leg metrics (course, TWA, sea angle, polar/boat speed, AWS/AWA)
    if leg.course_out is not None:
        style_centered_cell(ws.cell(r, 12), f"{leg.course_out:03d}°T")
        style_centered_cell(ws.cell(r, 13), leg.twa)

        # Sea From — color-coded by source tag (WD = forecast Wave Detail, SE = synoptic estimate)
        sea_from_fill = COLOR_GOOD_FILL if leg.sea_source_tag == "WD" else COLOR_NEUTRAL_FILL
        style_centered_cell(ws.cell(r, 14), leg.sea_from_label, fill_color=sea_from_fill)

        sea_angle_str = f"{leg.sea_angle}° {leg.sea_position}" if leg.sea_position else f"{leg.sea_angle}°"
        style_centered_cell(ws.cell(r, 15), sea_angle_str)

        style_number_cell(ws.cell(r, 16), leg.polar_speed, number_format="0.00")
        style_number_cell(ws.cell(r, 17), leg.boat_speed, number_format="0.00")
        style_number_cell(ws.cell(r, 18), leg.aws, number_format="0.0")
        style_centered_cell(ws.cell(r, 19), f"{leg.awa}°" if leg.awa else "—")
        style_text_cell(ws.cell(r, 20), leg.sail_mode)
    else:
        # Terminal WP — no leg metrics; render placeholders
        for c in (12, 13, 14, 15, 16, 17, 18, 19):
            style_centered_cell(ws.cell(r, c), "—")
        style_text_cell(ws.cell(r, 20), leg.sail_mode or "Inlet entry")

    # U: Rose is embedded separately via _embed_rose
    # V: Mini polar is embedded separately via _embed_mini_polar

    # W: Notes — wrapping text, top-aligned
    notes_cell = ws.cell(r, 23, value=leg.notes)
    notes_cell.alignment = align_text_top()
    notes_cell.font = Font(name="Calibri", size=10)
    notes_cell.border = thin_border()

    # X: Weather Risk — colored fill matches risk_color, wrapping text
    risk_cell = ws.cell(r, 24, value=leg.weather_risk)
    risk_cell.alignment = align_text_top()
    risk_cell.font = Font(name="Calibri", size=10)
    risk_cell.fill = fill(leg.risk_color)
    risk_cell.border = thin_border()

    # Y: Sea Source — italic gray
    seasrc_cell = ws.cell(r, 25, value=leg.sea_source)
    seasrc_cell.alignment = align_text_top()
    seasrc_cell.font = Font(name="Calibri", size=9, italic=True, color=COLOR_NOTE_FONT)
    seasrc_cell.border = thin_border()

    # Z, AA, AB: Cumulative time, sailing, motoring
    style_number_cell(ws.cell(r, 26), round(leg.cum_time_hr, 1), number_format="0.0")
    style_number_cell(ws.cell(r, 27), round(leg.cum_sailing_hr, 1), number_format="0.0")
    style_number_cell(ws.cell(r, 28), round(leg.cum_motoring_hr, 1), number_format="0.0")

    # === Row band — applied AFTER per-cell colors, only fills the cells
    # that don't already have meaningful coloring (ETA, Sea From, Weather
    # Risk all carry their own color). The band creates a subtle whole-row
    # tint for risk emphasis without overpowering the data cells.
    if leg.row_band:
        band_color = ROW_BAND_COLORS.get(leg.row_band)
        if band_color:
            protected_cols = {4, 14, 24}  # ETA, Sea From, Weather Risk keep their color
            for c in range(1, N_COLS + 1):
                if c in protected_cols:
                    continue
                cell = ws.cell(r, c)
                # Only apply band to cells without an explicit fill yet
                cur_rgb = cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else None
                if cur_rgb in (None, "00000000", "FFFFFFFF", "00FFFFFF"):
                    cell.fill = fill(band_color)

    # Row height — multi-line Notes/Risk/Sea Source columns need ~75-100px;
    # we set 95 to match the rose image height (120 px = ~95 Excel units).
    # The rose embed will re-set this for rows with a rose.
    ws.row_dimensions[r].height = 95


def _write_totals_row(ws, r, total_nm, total_time, sailing_total, motoring_total):
    """Write the PASSAGE TOTALS row with bold styling and subheader fill."""
    sail_pct = (sailing_total / total_time * 100) if total_time > 0 else 0
    motor_pct = (motoring_total / total_time * 100) if total_time > 0 else 0
    avg_spd = (total_nm / total_time) if total_time > 0 else 0

    label_cell = ws.cell(r, 2, value="PASSAGE TOTALS")
    label_cell.font = Font(name="Calibri", size=11, bold=True, color=COLOR_TITLE_FONT)
    label_cell.fill = fill(COLOR_SUBHEADER_FILL)
    label_cell.alignment = Alignment(horizontal="left", vertical="center")
    label_cell.border = thin_border()

    detail = (
        f"Total {total_time:.1f} hr   |   "
        f"Sailing {sailing_total:.1f} hr ({sail_pct:.0f}%)   |   "
        f"Motoring {motoring_total:.1f} hr ({motor_pct:.0f}%)   |   "
        f"Avg Speed {avg_spd:.1f} kt"
    )
    detail_cell = ws.cell(r, 3, value=detail)
    detail_cell.font = Font(name="Calibri", size=11, bold=True, color=COLOR_TITLE_FONT)
    detail_cell.fill = fill(COLOR_SUBHEADER_FILL)
    detail_cell.alignment = Alignment(horizontal="left", vertical="center")
    detail_cell.border = thin_border()
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=25)

    for col, val in [(26, total_time), (27, sailing_total), (28, motoring_total)]:
        c = ws.cell(r, col, value=round(val, 1))
        c.font = Font(name="Calibri", size=11, bold=True, color=COLOR_TITLE_FONT)
        c.fill = fill(COLOR_SUBHEADER_FILL)
        c.alignment = align_number()
        c.number_format = "0.0"
        c.border = thin_border()

    ws.row_dimensions[r].height = 24

    # WP column gets a blank styled cell to keep the row visually intact
    blank = ws.cell(r, 1, value="")
    blank.fill = fill(COLOR_SUBHEADER_FILL)
    blank.border = thin_border()


def _write_legend_block(ws, anchor_row):
    """Write the three color-key legends (ETA / Risk / Row band) + sea position glossary."""
    r = anchor_row

    # ETA color key
    ws.cell(r, 2, value="ETA color key").font = body_bold_font()
    r += 1
    _legend_row(ws, r,     "DAY",      COLOR_GOOD_FILL,    "Sunrise + 30 min to sunset – 30 min")
    _legend_row(ws, r + 1, "TWILIGHT", COLOR_NEUTRAL_FILL, "Civil dawn to sunrise + 30 min, or sunset – 30 to civil dusk")
    _legend_row(ws, r + 2, "NIGHT",    COLOR_BAD_FILL,     "After civil dusk before civil dawn (dark)")

    # Weather Risk key
    r += 4
    ws.cell(r, 2, value="Weather Risk color key").font = body_bold_font()
    r += 1
    _legend_row(ws, r,     "GREEN",  COLOR_GOOD_FILL,    "Benign conditions, no concerns")
    _legend_row(ws, r + 1, "YELLOW", COLOR_NEUTRAL_FILL, "Elevated risk — monitor (sporty seas, transitions)")
    _legend_row(ws, r + 2, "RED",    COLOR_BAD_FILL,     "Active hazard (frontal passage, severe convection, unsafe arrival)")

    # Row band key
    r += 4
    ws.cell(r, 2, value="Row band key (whole-row tint)").font = body_bold_font()
    r += 1
    _legend_row(ws, r,     "AMBER",     "FFE699", "Pre-frontal / pressure falling / front imminent")
    _legend_row(ws, r + 1, "YELLOW",    "FFEB9C", "High-attention departure into rough conditions")
    _legend_row(ws, r + 2, "SALMON",    "F8CBAD", "Tactical hazard (wind on the nose, etc.)")
    _legend_row(ws, r + 3, "LIGHT GRN", "E2EFDA", "Light-air recovery / favorable conditions")

    # Sea position glossary
    r += 5
    ws.cell(r, 2, value="Sea Position key (how the boat feels the wave)").font = body_bold_font()
    r += 1
    positions = [
        ("HD", "Head — within 30° of bow (pitching, slow)"),
        ("PB", "Port Bow — port side, 30-90° (forward of beam)"),
        ("PQ", "Port Quarter — port side, 90-150° (comfortable downwind)"),
        ("ST", "Stern — within 30° of stern (following, surf risk)"),
        ("SQ", "Starboard Quarter — starboard side, 90-150°"),
        ("SB", "Starboard Bow — starboard side, 30-90°"),
    ]
    for i, (code, desc) in enumerate(positions):
        code_cell = ws.cell(r + i, 3, value=code)
        code_cell.font = body_bold_font()
        code_cell.alignment = align_center()
        desc_cell = ws.cell(r + i, 4, value=desc)
        desc_cell.font = Font(name="Calibri", size=10)
        desc_cell.alignment = Alignment(horizontal="left", vertical="center")

    return r + len(positions)


def _write_footer_note(ws, r, passage):
    """Footer caption about polar / motor baseline. Uses vessel from passage."""
    vessel = (passage or {}).get("vessel", {}) if passage else {}
    vessel_label = vessel.get("designation", "vessel")
    design_label = vessel.get("design_number", "")
    motor_speed = vessel.get("motor_speed_kt", 6.0)
    polar_id = f"{vessel_label} {design_label} {vessel.get('load_config', 'half-load')}".strip()

    note_text = (
        f"Polar Speed (kt): pure bilinear interpolation of {polar_id} polar at "
        f"modeled (TWS, TWA) — reference value, no sea-state penalty applied. "
        f"Boat Speed (kt): polar × sea factor — the realistic value used to drive ETAs. "
        f"Motoring legs = {motor_speed} kt ({vessel_label} motor baseline)."
    )
    note = ws.cell(r, 2, value=note_text)
    note.font = caption_font()
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=20)
    ws.row_dimensions[r].height = 40


def _legend_row(ws, row, label, color, description):
    chip = ws.cell(row, 3, value=label)
    chip.fill = fill(color)
    chip.font = Font(name="Calibri", size=10, bold=True)
    chip.alignment = align_center()
    chip.border = thin_border()
    desc = ws.cell(row, 4, value=description)
    desc.font = Font(name="Calibri", size=10)
    desc.alignment = Alignment(horizontal="left", vertical="center")


# ======================================================================
# Wind/Sea rose embedding (column U)
# ======================================================================
def _embed_rose(ws, row, leg):
    """Render course-up wind/sea rose for this leg, embedded in column U."""
    if leg.course_out is None or not HAVE_CAIROSVG:
        cell = ws.cell(row, 21, value="—")
        cell.alignment = align_center()
        cell.border = thin_border()
        return

    try:
        png_buf = rose_png_bytes(
            course_deg=leg.course_out,
            wind_from_deg=leg.wind_dir_deg,
            sea_from_deg=leg.sea_from_deg,
            output_px=160,
        )
    except Exception:
        cell = ws.cell(row, 21, value="—")
        cell.alignment = align_center()
        cell.border = thin_border()
        return

    img = XLImage(png_buf)
    img.width = 120
    img.height = 120
    img.anchor = f"U{row}"
    ws.add_image(img)
    # Set row height ~95 to fit the 120 px image (Excel units ≈ 1.33 px each)
    ws.row_dimensions[row].height = 95


def _embed_timeline_strip(ws, anchor_row, legs):
    """Embed a wind/sea timeline chart spanning the passage.

    Anchors at column A and spans the full width of the data section.
    Provides at-a-glance view of where the rough patches are.
    """
    try:
        from ..charts import timeline_strip_png_bytes
        # Taller chart for readability — 340 px output → 320 px embed → ~17 rows of vertical space
        buf = timeline_strip_png_bytes(legs, output_w_px=1200, output_h_px=340)
        if buf is None:
            return
        img = XLImage(buf)
        img.width = 1100
        img.height = 320
        img.anchor = f"A{anchor_row}"
        ws.add_image(img)
        # Reserve ~17 rows × 20 units ≈ 340 px for the taller chart
        for r in range(anchor_row, anchor_row + 17):
            cur = ws.row_dimensions[r].height
            if cur is None or cur < 20:
                ws.row_dimensions[r].height = 20
    except Exception:
        # Silently degrade — the totals row + table still convey the data
        pass


def _embed_mini_polar(ws, row, leg, passage):
    """Render the small polar curve for this leg at its TWS, with a red dot
    marking the current TWA. Embedded in column V (22).

    Anchored right next to the wind/sea rose in column U so the two visuals
    sit side-by-side: rose tells you the geometry, polar tells you where
    you are on the speed curve.

    Skipped for terminal WPs (no leg metrics) or if the polar generator
    raises. Falls back to "—" in the cell in those cases.
    """
    from ..charts import mini_polar_png_bytes
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment

    # Skip if no outbound leg
    if leg.course_out is None or leg.twa is None:
        cell = ws.cell(row, 22, value="—")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
        return

    # Use the midpoint of the wind speed range as the TWS for the polar.
    # (The actual polar is bilinear-interpolated across TWS so any value
    # in the range gives a reasonable curve.)
    tws_mid = (leg.wind_kt_low + leg.wind_kt_high) / 2 if leg.wind_kt_low else 10

    # Get design number from the passage so we use the right polar grid
    # (D1170 for HR48, D1206 for HR54, etc.)
    design = (passage or {}).get("vessel", {}).get("design_number", "D1170")

    try:
        png_buf = mini_polar_png_bytes(
            tws=tws_mid,
            twa=leg.twa,
            design=design,
            output_px=160,
        )
    except Exception:
        cell = ws.cell(row, 22, value="—")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
        return

    img = XLImage(png_buf)
    img.width = 120
    img.height = 120
    img.anchor = f"V{row}"
    ws.add_image(img)
    # Row height already set to 95 by _embed_rose
