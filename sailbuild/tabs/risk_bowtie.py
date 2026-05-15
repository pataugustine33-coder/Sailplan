"""
Risk Bowtie tab — six-axis scoring of vessel against passage conditions.

Each axis maps a raw vessel/passage metric to a 1-10 score via linear
interpolation between two anchor values:
  - score 1 anchor = "worst" / concerning
  - score 10 anchor = "best" / exceptional

The scoring is driven by:
  - Static vessel metrics (RM, displacement, hull form) → Stability, Comfort,
    Capsize Resistance
  - Plan-specific metrics (peak wind, peak TWA, peak SOG required) → Speed
    Margin, Heel Margin
  - Static endurance (fuel range) vs total passage NM → Range/Endurance

Generated per-plan since Heel and Speed depend on actual plan conditions.
"""
from openpyxl.styles import Font, Alignment, PatternFill
from ..styles import style_header_cell, fill, COLOR_SUBHEADER_FILL, COLOR_GOOD_FILL


def render_risk_bowtie(ws, passage, plan, legs):
    """Render Risk Bowtie tab for one plan."""
    from openpyxl.drawing.image import Image as XLImage
    from ..charts import radar_chart_png_bytes

    vessel_label = passage.get("vessel", {}).get("designation", "Vessel")
    ws.cell(1, 1, value=f"{plan['tab_label']} — Risk Bowtie: {vessel_label} Comfort & Safety Radar").font = Font(size=14, bold=True)
    ws.merge_cells("A1:H1")

    ws.cell(2, 1, value=(
        f"Six-axis scoring (1-10, higher = better) for {vessel_label} against this plan's actual conditions. "
        "Higher score = better margin / comfort / safety. Scores 5-6 are typical for moderate cruisers in moderate conditions."
    )).alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 30

    # Compute scores
    scores = _compute_scores(passage, legs)

    # === Embed radar chart ===
    # The chart is the same data as the table below — but visual.
    # Polygon shape tells you the boat's profile at a glance:
    # balanced = good all-rounder, spiky = strengths/weaknesses, small = concerning.
    try:
        chart_buf = radar_chart_png_bytes(
            scores["axes"],
            title=f"{vessel_label} — {plan['tab_label']} Risk Profile",
            output_px=520,
        )
        img = XLImage(chart_buf)
        img.width = 480
        img.height = 480
        img.anchor = "G4"
        ws.add_image(img)
        # Reserve enough vertical space for the chart (~26 rows × ~18 units ≈ chart height)
        for r in range(4, 30):
            if ws.row_dimensions[r].height is None or ws.row_dimensions[r].height < 20:
                ws.row_dimensions[r].height = 20
    except Exception:
        pass  # If chart generation fails, fall through to table-only view

    # Headers
    headers = ["Dimension", "Score (1-10)", "Score 1 anchor", "Score 10 anchor", "What it measures", "Raw value (this plan)"]
    for i, h in enumerate(headers, 1):
        style_header_cell(ws.cell(4, i, value=h))

    # Score rows
    for i, s in enumerate(scores["axes"]):
        r = 5 + i
        ws.cell(r, 1, value=s["name"]).font = Font(bold=True)
        score_cell = ws.cell(r, 2, value=round(s["score"], 1))
        # Color-code the score cell
        if s["score"] >= 7:
            score_cell.fill = fill(COLOR_GOOD_FILL)
        elif s["score"] >= 4:
            score_cell.fill = fill("FFEB9C")
        else:
            score_cell.fill = fill("FFC7CE")
        score_cell.font = Font(bold=True, size=12)
        score_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r, 3, value=s["anchor_low"])
        ws.cell(r, 4, value=s["anchor_high"])
        ws.cell(r, 5, value=s["description"]).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(r, 6, value=s["raw"])

    # Average row
    avg_row = 5 + len(scores["axes"])
    ws.cell(avg_row, 1, value="AVERAGE (overall fit)").font = Font(bold=True)
    avg_cell = ws.cell(avg_row, 2, value=round(scores["average"], 2))
    avg_cell.font = Font(bold=True, size=14)
    avg_cell.fill = fill(COLOR_SUBHEADER_FILL)
    avg_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Profile reading narrative
    pr_row = avg_row + 2
    ws.cell(pr_row, 1, value="Profile Reading").font = Font(bold=True)
    narrative = _profile_narrative(scores, plan)
    nar_cell = ws.cell(pr_row + 1, 1, value=narrative)
    nar_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=pr_row + 1, start_column=1, end_row=pr_row + 1, end_column=6)
    ws.row_dimensions[pr_row + 1].height = 180

    # Methodology block
    meth_row = pr_row + 3
    ws.cell(meth_row, 1, value="Methodology notes").font = Font(bold=True)
    meth_text = (
        "SCORING: Each axis maps a raw metric to a 1-10 score using linear interpolation between two anchor "
        "values (column C = score-1 threshold, column D = score-10 threshold). Anchors are calibrated to the "
        "cruising-yacht range — 5 represents a typical moderate cruiser, 8+ is exceptional, <4 means concern.\n\n"
        "STATIC vs DYNAMIC: Stability, Motion Comfort, Capsize Resistance, and Range are STATIC metrics — they "
        "don't change with forecast. Speed Margin and Heel Margin are DYNAMIC — recomputed from the plan's "
        "actual peak conditions. So when the forecast changes, the dynamic scores may shift while the static "
        "ones stay constant.\n\n"
        "LIMITATIONS: This is a heuristic scorecard for go/no-go discussion, not a quantitative seaworthiness "
        "rating. The numerical anchors come from cruising literature (Brewer Comfort Ratio, CSF, etc.) and "
        "project judgment, not from regulatory or class standards."
    )
    mtxt_cell = ws.cell(meth_row + 1, 1, value=meth_text)
    mtxt_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=meth_row + 1, start_column=1, end_row=meth_row + 1, end_column=6)
    ws.row_dimensions[meth_row + 1].height = 180

    # Column widths
    for col, w in [("A", 24), ("B", 14), ("C", 22), ("D", 22), ("E", 44), ("F", 22)]:
        ws.column_dimensions[col].width = w


def _compute_scores(passage, legs):
    """Compute the six axis scores from vessel metrics + plan legs."""
    vessel = passage["vessel"]
    total_nm = passage["passage"]["total_nm"]

    # === Plan-specific dynamic metrics ===
    sailing_legs = [l for l in legs if l.course_out is not None and not l.is_motor]
    all_legs = [l for l in legs if l.course_out is not None]

    peak_wind_kt = max((l.wind_kt_high for l in all_legs), default=0)
    peak_polar = max((l.polar_speed for l in sailing_legs), default=0)
    peak_aws = max((l.aws for l in sailing_legs), default=0)

    # Required average SOG (for passage to complete in time available)
    total_time = legs[-1].cum_time_hr if legs else 1
    avg_required_sog = total_nm / total_time if total_time > 0 else 0

    # Peak heel estimate: function of (wind, sail config, TWA).
    # Heuristic: heel ~ 2° per kt AWS for close-hauled, ~1° per kt AWS for broad reach
    # For HR 48 at typical loads.
    peak_heel_deg = 0
    for leg in sailing_legs:
        if leg.twa < 60:
            heel = leg.aws * 2.0
        elif leg.twa < 90:
            heel = leg.aws * 1.5
        elif leg.twa < 120:
            heel = leg.aws * 1.0
        else:
            heel = leg.aws * 0.7
        peak_heel_deg = max(peak_heel_deg, heel)

    # === Static metrics ===
    # Convert RM @ 30° from kg·m to lb·ft for the Stability axis (1 kg·m = 7.233 lb·ft)
    rm_30_lbft = vessel["rm_30deg_kgm"] * 7.233
    # Motor range
    motor_range_nm = (vessel["fuel_capacity_gal"] / vessel["fuel_burn_gph"]) * vessel["motor_speed_kt"]
    # Brewer Comfort Ratio and Capsize Screening Formula — pulled from vessel
    # YAML so each design (HR 48, HR 54, etc.) gets its own values.
    # Defaults to HR 48 values for backward compatibility with older YAMLs.
    brewer_cr = vessel.get("brewer_cr", 42.9)
    csf = vessel.get("csf", 1.66)

    axes = [
        {
            "name": "Stability",
            "score": _linear_score(rm_30_lbft, 50000, 150000),
            "raw": f"{int(rm_30_lbft):,} lb·ft",
            "anchor_low": "RM @ 30° = 50,000 lb·ft",
            "anchor_high": "RM @ 30° = 150,000 lb·ft",
            "description": "Working-range righting moment at 30° heel. Indicates resistance to capsizing forces in normal sailing conditions.",
        },
        {
            "name": "Motion Comfort",
            "score": _linear_score(brewer_cr, 25, 60),
            "raw": f"Brewer CR = {brewer_cr}",
            "anchor_low": "Brewer CR = 25 (racer)",
            "anchor_high": "Brewer CR = 60 (heavy cruiser)",
            "description": "Brewer Comfort Ratio — motion in seas. Higher = more comfortable, slower motion.",
        },
        {
            "name": "Capsize Resistance",
            "score": _linear_score_inverted(csf, 2.0, 1.4),
            "raw": f"CSF = {csf}",
            "anchor_low": "CSF = 2.0 (concerning)",
            "anchor_high": "CSF = 1.4 (excellent)",
            "description": "Capsize Screening Formula. LOWER is safer. <2.0 acceptable for offshore, <1.8 preferred.",
        },
        {
            "name": "Speed Margin",
            "score": _linear_score(peak_polar / avg_required_sog if avg_required_sog else 1.0, 1.0, 1.5),
            "raw": f"{peak_polar/avg_required_sog:.2f} ({peak_polar:.2f}/{avg_required_sog:.2f})" if avg_required_sog else "N/A",
            "anchor_low": "Polar / Required = 1.0 (no margin)",
            "anchor_high": "Polar / Required = 1.5 (50% margin)",
            "description": "Peak polar speed ÷ passage avg required speed. Headroom for slow-leg recovery and contingencies.",
        },
        {
            "name": "Heel Margin",
            "score": _linear_score_inverted(peak_heel_deg, 25, 10),
            "raw": f"~{peak_heel_deg:.0f}° peak heel",
            "anchor_low": "Peak heel = 25° (reef trigger)",
            "anchor_high": "Peak heel = 10° (comfortable)",
            "description": "Estimated peak heel angle vs 25° reefing trigger. Higher score = less heel, more comfortable single-handed.",
        },
        {
            "name": "Range / Endurance",
            "score": _linear_score(motor_range_nm, 500, 1100),
            "raw": f"{int(motor_range_nm)} NM motor range",
            "anchor_low": "Motor range = 500 NM",
            "anchor_high": "Motor range = 1100 NM",
            "description": "Diesel range at 6 kt cruise. Substantial reserve enables contingency motoring for becalmed or weather-avoidance.",
        },
    ]

    average = sum(a["score"] for a in axes) / len(axes)

    return {"axes": axes, "average": average, "context": {
        "peak_wind_kt": peak_wind_kt,
        "peak_polar": peak_polar,
        "peak_heel_deg": peak_heel_deg,
        "avg_required_sog": avg_required_sog,
        "motor_range_nm": motor_range_nm,
    }}


def _linear_score(raw, anchor_low, anchor_high):
    """Map raw to 1-10 score with anchor_low → 1 and anchor_high → 10."""
    if anchor_high == anchor_low:
        return 5.5
    score = 1 + 9 * (raw - anchor_low) / (anchor_high - anchor_low)
    return max(1.0, min(10.0, score))


def _linear_score_inverted(raw, anchor_low, anchor_high):
    """Inverted: anchor_low → 1, anchor_high → 10 where anchor_low > anchor_high (lower raw = better)."""
    if anchor_high == anchor_low:
        return 5.5
    score = 1 + 9 * (anchor_low - raw) / (anchor_low - anchor_high)
    return max(1.0, min(10.0, score))


def _profile_narrative(scores, plan):
    """Generate natural-language profile reading."""
    axes_by_score = sorted(scores["axes"], key=lambda a: a["score"], reverse=True)
    strongest = axes_by_score[0]
    weakest = axes_by_score[-1]
    ctx = scores["context"]

    lines = [f"AVERAGE SCORE: {scores['average']:.2f}/10 ({plan['tab_label']})", ""]
    lines.append(f"  • Strongest axis: {strongest['name']} ({strongest['score']:.1f}) — {strongest['raw']}.")
    lines.append(f"  • Weakest axis: {weakest['name']} ({weakest['score']:.1f}) — {weakest['raw']}.")
    lines.append("")
    lines.append(
        f"Plan-specific conditions driving dynamic scores: peak wind {ctx['peak_wind_kt']:.0f} kt, "
        f"peak polar speed {ctx['peak_polar']:.1f} kt, estimated peak heel {ctx['peak_heel_deg']:.0f}°, "
        f"required average SOG {ctx['avg_required_sog']:.1f} kt to complete passage."
    )
    lines.append("")
    lines.append("Score interpretation:")
    lines.append("  • 8-10: Exceptional — boat well-matched to conditions, large margins.")
    lines.append("  • 5-7:  Acceptable — moderate cruiser in moderate conditions, expected range.")
    lines.append("  • <5:   Concern — re-evaluate before departure; consider delay or alternate plan.")
    return "\n".join(lines)
