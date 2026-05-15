"""
Generators for all supporting tabs: Format Reference, Waypoints, Vessel
Particulars, Live Buoy Data, Refresh Cadence, Forecast Sources, Buoys by WP,
Forecast Products, URL Quick Reference, Glossary, Verification Scorecard.

Each is a single function `render_<name>(ws, ...)` that writes the tab.
"""
from datetime import datetime, timezone, timedelta
import re
from openpyxl.styles import Font, Alignment, PatternFill
from ..styles import (
    style_header_cell, fill,
    COLOR_GOOD_FILL, COLOR_NEUTRAL_FILL, COLOR_BAD_FILL, COLOR_SUBHEADER_FILL,
    COLOR_DANGER_FONT,
)


# ==============================================================
# Forecast freshness helpers
# ==============================================================
# NWS marine forecasts (CWF, AFD) issue ~4× daily on a 6 hr cycle. Tier thresholds:
#   FRESH    < 12 hr   within current or previous cycle, operational
#   STALE    12–72 hr  multiple cycles late, refresh recommended
#   ARCHIVED > 72 hr   not operational, do not use as live forecast
FRESHNESS_TIERS = [
    (12,  "FRESH",    COLOR_GOOD_FILL),     # green
    (72,  "STALE",    COLOR_NEUTRAL_FILL),  # yellow
    (None, "ARCHIVED", COLOR_BAD_FILL),     # red
]

# Month abbrev → number, for parsing NWS-style "...May 12 2026" labels
_MONTH = {"Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
          "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12}

# US timezone abbrev → UTC offset hours
_TZ_OFFSET = {"EDT": -4, "EST": -5, "CDT": -5, "CST": -6, "MDT": -6, "MST": -7,
              "PDT": -7, "PST": -8, "AKDT": -8, "AKST": -9, "HDT": -9, "HST": -10}


def parse_issued_datetime(value):
    """Best-effort parse of an `issued` field into an aware UTC datetime.

    Accepts ISO 8601 ('2026-05-12T11:18:00-04:00') OR raw NWS label
    ('1118 AM EDT Tue May 12 2026'). Returns None if unparseable.
    """
    if not value:
        return None
    s = str(value).strip()

    # Try ISO 8601 first
    try:
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except (ValueError, TypeError):
        pass

    # Fall back to NWS bulletin format: "1118 AM EDT Tue May 12 2026"
    # parts: [0]=time, [1]=AM/PM, [2]=tz, [3]=dow, [4]=month, [5]=day, [6]=year
    parts = s.split()
    if len(parts) < 7:
        return None
    try:
        time_str, ampm, tz_abbr = parts[0], parts[1].upper(), parts[2].upper()
        month_name, day_str, year_str = parts[4], parts[5], parts[6]
        # time_str is "HMM" or "HHMM" (no separator). Last two chars = minutes.
        if len(time_str) < 3:
            return None
        m = int(time_str[-2:])
        h = int(time_str[:-2])
        if ampm == "PM" and h != 12:
            h += 12
        elif ampm == "AM" and h == 12:
            h = 0
        month = _MONTH.get(month_name)
        if month is None:
            return None
        offset_h = _TZ_OFFSET.get(tz_abbr, 0)
        local = datetime(int(year_str), month, int(day_str), h, m,
                         tzinfo=timezone(timedelta(hours=offset_h)))
        return local.astimezone(timezone.utc)
    except (ValueError, KeyError, IndexError):
        return None


def classify_freshness(issued_value, now_utc=None):
    """Classify a forecast's freshness. Returns (age_string, status, fill_hex).

    Returns ("?", "UNKNOWN", None) if the issued field can't be parsed.
    """
    if now_utc is None:
        now_utc = datetime.now(timezone.utc)
    dt = parse_issued_datetime(issued_value)
    if dt is None:
        return ("?", "UNKNOWN", None)
    age_hr = (now_utc - dt).total_seconds() / 3600.0
    if age_hr < 0:
        age_hr = 0  # clock skew: treat future timestamps as just-issued
    # Format age compactly
    if age_hr < 24:
        age_str = f"{age_hr:.1f} hr"
    elif age_hr < 14 * 24:
        age_str = f"{age_hr / 24:.1f} days"
    else:
        age_str = f"{age_hr / (24 * 7):.1f} weeks"
    # Pick tier
    for limit, label, color in FRESHNESS_TIERS:
        if limit is None or age_hr < limit:
            return (age_str, label, color)
    return (age_str, "ARCHIVED", COLOR_BAD_FILL)


# ==============================================================
# Waypoints
# ==============================================================
def render_waypoints(ws, passage, total_nm):
    """Waypoints tab — route geographic detail with paste-ready CSV block."""
    from ..styles import (
        style_page_title, style_page_subtitle, style_section_header,
        style_table_header, style_number_cell, style_centered_cell, style_text_cell,
        body_bold_font, caption_font, set_column_widths,
    )

    # === Title ===
    style_page_title(
        ws.cell(1, 1),
        f"Route Waypoints — {passage['passage']['origin']} → {passage['passage']['destination']}",
    )
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 26

    style_page_subtitle(
        ws.cell(2, 1),
        f"Total distance {total_nm:.1f} NM across {len(passage['waypoints'])} waypoints. "
        f"Lat/Lon are decimal degrees (positive N, positive W). Paste-ready coords are "
        f"in the format Google My Maps / OpenCPN expect.",
    )
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 30

    # === Headers (row 4) ===
    headers = ["WP", "Description", "Lat (°N)", "Lon (°W)",
               "Map Coords (paste-ready)", "Cum NM", "Leg NM", "Course (°T)"]
    for i, h in enumerate(headers, 1):
        style_table_header(ws.cell(4, i, value=h))
    ws.row_dimensions[4].height = 32

    # === Data rows (row 5+) ===
    wps = passage["waypoints"]
    for i, wp in enumerate(wps):
        r = 5 + i
        lat = wp["lat"]
        lon = abs(wp["lon"])
        leg_nm = round(wp["cum_nm"] - wps[i-1]["cum_nm"], 1) if i > 0 else 0

        style_centered_cell(ws.cell(r, 1), wp["id"], bold=True)
        style_text_cell(ws.cell(r, 2), wp["name"], wrap=False)
        style_number_cell(ws.cell(r, 3), lat, number_format="0.0000")
        style_number_cell(ws.cell(r, 4), lon, number_format="0.0000")
        style_centered_cell(ws.cell(r, 5), f"{lat:.4f}, -{lon:.4f}")
        style_number_cell(ws.cell(r, 6), wp["cum_nm"], number_format="0.0")
        style_number_cell(ws.cell(r, 7), leg_nm, number_format="0.0")
        course_out = wp.get("course_out")
        style_centered_cell(
            ws.cell(r, 8),
            f"{int(course_out):03d}°T" if course_out is not None else "—",
        )
        ws.row_dimensions[r].height = 22

    # === CSV block (paste into Google My Maps / OpenCPN) ===
    csv_row = 5 + len(wps) + 3
    style_section_header(ws.cell(csv_row, 1), "Mapping CSV — paste into Google My Maps or OpenCPN")
    ws.merge_cells(start_row=csv_row, start_column=1, end_row=csv_row, end_column=8)
    ws.row_dimensions[csv_row].height = 22

    csv_headers = ["Name", "Latitude", "Longitude", "Description"]
    for i, h in enumerate(csv_headers, 1):
        style_table_header(ws.cell(csv_row + 1, i, value=h))
    for i, wp in enumerate(wps):
        r = csv_row + 2 + i
        style_text_cell(ws.cell(r, 1), wp["id"], wrap=False)
        style_number_cell(ws.cell(r, 2), wp["lat"], number_format="0.0000")
        style_number_cell(ws.cell(r, 3), -abs(wp["lon"]), number_format="0.0000")
        style_text_cell(ws.cell(r, 4), wp["name"], wrap=False)

    # === Map links + total ===
    coords = "/".join(f"{wp['lat']},-{abs(wp['lon'])}" for wp in wps)
    link_row = csv_row + 2 + len(wps) + 2

    style_section_header(ws.cell(link_row, 1), "Quick links")
    ws.merge_cells(start_row=link_row, start_column=1, end_row=link_row, end_column=8)

    ws.cell(link_row + 1, 1, value="Google Maps directions").font = body_bold_font()
    url_cell = ws.cell(link_row + 1, 2, value=f"https://www.google.com/maps/dir/{coords}")
    url_cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
    ws.merge_cells(start_row=link_row + 1, start_column=2, end_row=link_row + 1, end_column=8)

    total_cell = ws.cell(link_row + 3, 1,
                         value=f"Total: {total_nm:.1f} NM   |   {len(wps)} waypoints")
    total_cell.font = body_bold_font()

    # === Column widths ===
    set_column_widths(ws, {
        "A": 6, "B": 32, "C": 12, "D": 12, "E": 26, "F": 10, "G": 10, "H": 14,
    })

    # Freeze panes so header stays visible
    ws.freeze_panes = "A5"


# ==============================================================
# Vessel Particulars
# ==============================================================
def render_vessel_particulars(ws, passage):
    """Vessel Particulars — polar grid + boat data sheet."""
    from ..styles import (
        style_page_title, style_page_subtitle, style_section_header,
        style_table_header, style_number_cell, style_centered_cell, style_text_cell,
        body_bold_font, set_column_widths,
    )

    v = passage["vessel"]

    # === Title ===
    style_page_title(
        ws.cell(1, 1),
        f"{v['designation']} ({v['design_number']}) — Vessel Particulars",
    )
    ws.merge_cells("A1:K1")
    ws.row_dimensions[1].height = 26

    style_page_subtitle(
        ws.cell(2, 1),
        f"Frers-designed Hallberg-Rassy {v['load_config']}. "
        f"Polar from Frers VPP. All metrics from manufacturer data sheets and "
        f"stability documents.",
    )
    ws.merge_cells("A2:K2")
    ws.row_dimensions[2].height = 28

    # === Polar grid ===
    from ..polar import TWS_GRID, TWA_GRID, POLARS
    design_number = v.get("design_number", "D1170")
    vs_grid = POLARS.get(design_number, POLARS["D1170"])

    style_section_header(
        ws.cell(4, 1),
        f"Polar — Boat speed (kt) at TWA × TWS — {design_number} {v['load_config']} (Frers VPP)",
    )
    ws.merge_cells("A4:J4")
    ws.row_dimensions[4].height = 22

    # Header row: TWA label + TWS columns
    style_table_header(ws.cell(5, 1, value="TWA (°)"))
    for i, tws in enumerate(TWS_GRID, 2):
        style_table_header(ws.cell(5, i, value=f"TWS {tws}"))
    ws.row_dimensions[5].height = 28

    # Data rows
    for i, twa in enumerate(TWA_GRID):
        r = 6 + i
        style_centered_cell(ws.cell(r, 1), int(twa), bold=True)
        for j, tws in enumerate(TWS_GRID, 2):
            style_number_cell(ws.cell(r, j), float(vs_grid[i, j-2]), number_format="0.00")
        ws.row_dimensions[r].height = 20

    # === Particulars table ===
    table_row = 6 + len(TWA_GRID) + 2
    style_section_header(ws.cell(table_row, 1), "Vessel Particulars & Specifications")
    ws.merge_cells(start_row=table_row, start_column=1, end_row=table_row, end_column=3)
    ws.row_dimensions[table_row].height = 22

    for i, h in enumerate(["Parameter", "Value", "Comment / Source"], 1):
        style_table_header(ws.cell(table_row + 1, i, value=h))
    ws.row_dimensions[table_row + 1].height = 28

    rows = [
        ("Designer",                          "Germán Frers", ""),
        ("CE category",                       "A", "Ocean — unlimited offshore"),
        ("Designation",                       v["designation"], ""),
        ("Design number",                     v["design_number"], ""),
        ("Load configuration",                v["load_config"], ""),
        ("Empty displacement",                f"{v['displacement_lb']:,} lb", "From Hallberg-Rassy data sheet"),
        ("Motor speed (cruise)",              f"{v['motor_speed_kt']} kt", "Diesel at moderate RPM"),
        ("Motor crossover TWS",               f"{v['motor_crossover_tws_kt']} kt", "Below this, motor faster than sailing"),
        ("Fuel capacity",                     f"{v['fuel_capacity_gal']} gal", ""),
        ("Fuel burn",                         f"{v['fuel_burn_gph']} gph", "At cruise"),
        ("Motor range",                       f"~{int(v['fuel_capacity_gal']/v['fuel_burn_gph']*v['motor_speed_kt'])} NM",
                                              "Tank ÷ burn × motor speed"),
        ("RMC (righting moment coeff.)",      f"{v['rmc_kgm']} kg·m", ""),
        ("RM @ 30°",                          f"{v['rm_30deg_kgm']:,} kg·m", "Righting moment at 30° heel"),
        ("RM max",                            f"{v['rm_max_kgm']:,} kg·m", "Peak righting moment"),
    ]
    if "brewer_cr" in v:
        rows.append(("Brewer Comfort Ratio", f"{v['brewer_cr']:.1f}",
                    ">50 = heavy cruiser, 40-50 = moderate"))
    if "csf" in v:
        rows.append(("Capsize Screening Formula", f"{v['csf']:.2f}",
                    "<2.0 acceptable offshore; lower = safer"))

    for i, (param, val, comment) in enumerate(rows):
        r = table_row + 2 + i
        style_text_cell(ws.cell(r, 1), param, wrap=False)
        ws.cell(r, 1).font = body_bold_font()
        style_text_cell(ws.cell(r, 2), val, wrap=False)
        ws.cell(r, 2).alignment = Alignment(horizontal="right", vertical="center")
        style_text_cell(ws.cell(r, 3), comment, wrap=True)
        ws.row_dimensions[r].height = 20

    # === Column widths ===
    set_column_widths(ws, {
        "A": 32, "B": 22, "C": 38,
        "D": 10, "E": 10, "F": 10, "G": 10, "H": 10, "I": 10, "J": 10,
    })

    ws.freeze_panes = "A5"


# ==============================================================
# Live Buoy Data
# ==============================================================
def render_live_buoy_data(ws, buoys):
    """Live Buoy Data tab — current NDBC observations + pressure tracking + findings."""
    from ..styles import (
        style_page_title, style_page_subtitle, style_section_header,
        style_table_header, style_number_cell, style_centered_cell, style_text_cell,
        body_bold_font, caption_font, set_column_widths,
    )

    style_page_title(
        ws.cell(1, 1),
        f"Live Buoy Data — NDBC realtime2 ground truth",
    )
    ws.merge_cells("A1:I1")
    ws.row_dimensions[1].height = 26

    style_page_subtitle(
        ws.cell(2, 1),
        f"Latest pull: {buoys['pull']['label']}.   "
        f"Live buoy data must be <2 hours old at every checkpoint — "
        f"if older than 2 hours, re-pull from realtime2 before any GO/NO-GO decision.",
    )
    ws.merge_cells("A2:I2")
    ws.row_dimensions[2].height = 30

    # === Section 1: Latest observations ===
    style_section_header(ws.cell(4, 1), "Latest observations (most recent reading per station)")
    ws.merge_cells("A4:I4")
    ws.row_dimensions[4].height = 22

    headers = ["Station", "Location", "Reading time", "Wind dir",
               "Wind (kt)", "Gust (kt)", "Pressure (inHg)", "Trend", "Status"]
    for i, h in enumerate(headers, 1):
        style_table_header(ws.cell(5, i, value=h))
    ws.row_dimensions[5].height = 32

    for i, st in enumerate(buoys["stations"]):
        r = 6 + i
        is_offline = st.get("status") == "offline"

        style_centered_cell(ws.cell(r, 1), str(st["id"]), bold=True)
        style_text_cell(ws.cell(r, 2), st["name"], wrap=True)

        if is_offline:
            style_centered_cell(
                ws.cell(r, 3),
                f"OFFLINE since {st.get('offline_since', '')}",
                fill_color=COLOR_BAD_FILL,
            )
            for col in (4, 5, 6, 7, 8):
                style_centered_cell(ws.cell(r, col), "—")
            style_centered_cell(ws.cell(r, 9), "OFFLINE", fill_color=COLOR_BAD_FILL)
        else:
            style_centered_cell(ws.cell(r, 3), st.get("reading_time", ""))
            wind_dir = st.get("wind_dir") or "—"
            wind_deg = st.get("wind_dir_deg")
            wind_dir_str = f"{wind_deg:03d}° {wind_dir}" if wind_deg else wind_dir
            style_centered_cell(ws.cell(r, 4), wind_dir_str)

            wind_kt = st.get("wind_kt")
            style_number_cell(ws.cell(r, 5), wind_kt, number_format="0.0") if wind_kt is not None \
                else style_centered_cell(ws.cell(r, 5), "—")
            gust_kt = st.get("gust_kt")
            style_number_cell(ws.cell(r, 6), gust_kt, number_format="0.0") if gust_kt is not None \
                else style_centered_cell(ws.cell(r, 6), "—")

            pressure = st.get("pressure_hpa") or st.get("pressure_inhg")
            if pressure and pressure > 100:  # hPa
                inhg = pressure * 0.02953
                style_number_cell(ws.cell(r, 7), inhg, number_format="0.00")
            elif pressure:
                style_number_cell(ws.cell(r, 7), pressure, number_format="0.00")
            else:
                style_centered_cell(ws.cell(r, 7), "—")

            trend = st.get("pressure_trend", "—")
            trend_lower = trend.lower() if isinstance(trend, str) else ""
            trend_fc = None
            if "rising" in trend_lower or "steady" in trend_lower or "level" in trend_lower:
                trend_fc = COLOR_GOOD_FILL
            elif "falling fast" in trend_lower or "bottoming" in trend_lower:
                trend_fc = COLOR_BAD_FILL
            elif "falling" in trend_lower:
                trend_fc = COLOR_NEUTRAL_FILL
            style_text_cell(ws.cell(r, 8), trend, fill_color=trend_fc)

            status_text = st.get("status", "fresh")
            status_fc = COLOR_GOOD_FILL if status_text == "fresh" else COLOR_NEUTRAL_FILL
            style_centered_cell(ws.cell(r, 9), status_text.upper(), fill_color=status_fc, bold=True)

        ws.row_dimensions[r].height = 22

    # === Section 2: Pressure trend tracking ===
    pt_row = 6 + len(buoys["stations"]) + 2
    style_section_header(ws.cell(pt_row, 1), "Pressure trend tracking — primary front-timing indicator")
    ws.merge_cells(start_row=pt_row, start_column=1, end_row=pt_row, end_column=9)
    ws.row_dimensions[pt_row].height = 22

    headers = ["Station", "Time", "Pressure (inHg)", "Δ vs prior", "Interpretation"]
    for i, h in enumerate(headers, 1):
        style_table_header(ws.cell(pt_row + 1, i, value=h))
    ws.row_dimensions[pt_row + 1].height = 32

    for i, p in enumerate(buoys.get("pressure_trend", [])):
        r = pt_row + 2 + i
        style_centered_cell(ws.cell(r, 1), p["station"], bold=True)
        style_centered_cell(ws.cell(r, 2), p["time"])
        style_number_cell(ws.cell(r, 3), p["pressure"], number_format="0.00")
        style_centered_cell(ws.cell(r, 4), p["delta"])
        style_text_cell(ws.cell(r, 5), p["interpretation"])
        # Merge interpretation across remaining columns for readability
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=9)
        ws.row_dimensions[r].height = 45

    # === Section 3: Findings + front passage narrative ===
    findings_row = pt_row + 2 + len(buoys.get("pressure_trend", [])) + 2

    if "findings" in buoys and buoys["findings"]:
        style_section_header(ws.cell(findings_row, 1), "Verification findings")
        ws.merge_cells(start_row=findings_row, start_column=1,
                       end_row=findings_row, end_column=9)
        ws.row_dimensions[findings_row].height = 22

        f_cell = ws.cell(findings_row + 1, 1, value=buoys["findings"])
        f_cell.alignment = Alignment(wrap_text=True, vertical="top")
        f_cell.font = Font(name="Calibri", size=10)
        ws.merge_cells(start_row=findings_row + 1, start_column=1,
                       end_row=findings_row + 1, end_column=9)
        ws.row_dimensions[findings_row + 1].height = 100
        next_row = findings_row + 3
    else:
        next_row = findings_row

    if "front_passage" in buoys:
        style_section_header(ws.cell(next_row, 1), "Front passage narrative")
        ws.merge_cells(start_row=next_row, start_column=1,
                       end_row=next_row, end_column=9)
        ws.row_dimensions[next_row].height = 22

        fp_cell = ws.cell(next_row + 1, 1, value=buoys["front_passage"]["description"])
        fp_cell.alignment = Alignment(wrap_text=True, vertical="top")
        fp_cell.font = Font(name="Calibri", size=10)
        ws.merge_cells(start_row=next_row + 1, start_column=1,
                       end_row=next_row + 1, end_column=9)
        ws.row_dimensions[next_row + 1].height = 130

    # === Column widths ===
    set_column_widths(ws, {
        "A": 10, "B": 32, "C": 22, "D": 16,
        "E": 10, "F": 10, "G": 14, "H": 32, "I": 12,
    })

    # Freeze panes after header section
    ws.freeze_panes = "A6"


# ==============================================================
# Forecast Sources by WP
# ==============================================================
def render_forecast_sources(ws, passage, forecast):
    """Forecast Sources by WP — attribution trail showing which NWS zone covers each WP."""
    from ..styles import (
        style_page_title, style_page_subtitle, style_section_header,
        style_table_header, style_centered_cell, style_text_cell,
        body_bold_font, set_column_widths, fill, thin_border, COLOR_BAND_LIGHT,
    )

    cycle = forecast["cycle"]

    style_page_title(ws.cell(1, 1), "NWS Forecast Sources by Waypoint")
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 26

    cwf_bits = [f"CWF{k[3:].upper()}: {v['label_full']}"
                for k, v in cycle.items() if k.startswith("cwf") and isinstance(v, dict)]
    afd_bits = [f"AFD{k[3:].upper()}: {v['label_full']}"
                for k, v in cycle.items() if k.startswith("afd") and isinstance(v, dict)]
    cycle_summary = "   |   ".join(cwf_bits + afd_bits)

    style_page_subtitle(
        ws.cell(2, 1),
        f"Attribution trail showing which NWS forecast zone covers each waypoint. "
        f"Cycle in use: {cycle_summary}",
    )
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 38

    # === Section 1: WP-to-zone mapping ===
    style_section_header(ws.cell(4, 1), "Waypoint → zone assignment (this passage)")
    ws.merge_cells("A4:E4")
    ws.row_dimensions[4].height = 22

    headers = ["WP", "Position (lat/lon)", "NWS Office", "Zone", "Zone Description"]
    for i, h in enumerate(headers, 1):
        style_table_header(ws.cell(5, i, value=h))
    ws.row_dimensions[5].height = 28

    # Build zone lookup
    zone_lookup = {}
    office_long_names = {
        "CHS": "NWS Charleston SC (KCHS)",
        "ILM": "NWS Wilmington NC (KILM)",
        "MHX": "NWS Newport/Morehead City NC (KMHX)",
        "AKQ": "NWS Wakefield VA (KAKQ)",
        "JAX": "NWS Jacksonville FL (KJAX)",
        "MFL": "NWS Miami FL (KMFL)",
        "KEY": "NWS Key West FL (KKEY)",
        "BOX": "NWS Boston MA (KBOX)",
    }
    for zone_id, zone_data in forecast.get("zones", {}).items():
        if isinstance(zone_data, dict):
            office_code = zone_data.get("office", "")
            office_label = office_long_names.get(office_code, office_code)
            desc = zone_data.get("description", "")
            zone_lookup[zone_id] = (office_label, desc)
    # Passage YAML overrides
    for office_info in passage.get("forecast_offices", []):
        for zone, desc in office_info["zones"].items():
            zone_lookup[zone] = (office_info["office"], desc)

    first_plan_id = passage["plans"][0]["id"]
    assignments = forecast["waypoint_assignments"].get(first_plan_id, {})
    for i, wp in enumerate(passage["waypoints"]):
        r = 6 + i
        band = COLOR_BAND_LIGHT if i % 2 == 1 else None
        zone = assignments.get(wp["id"], {}).get("zone", "—")
        office, zone_desc = zone_lookup.get(zone, ("—", "—"))

        style_centered_cell(ws.cell(r, 1), wp["id"], fill_color=band, bold=True)
        style_centered_cell(ws.cell(r, 2),
                            f"{wp['lat']:.4f}°N  {abs(wp['lon']):.4f}°W",
                            fill_color=band)
        style_text_cell(ws.cell(r, 3), office, fill_color=band, wrap=False)
        style_centered_cell(ws.cell(r, 4), zone, fill_color=band, bold=True)
        style_text_cell(ws.cell(r, 5), zone_desc, fill_color=band)
        ws.row_dimensions[r].height = 22

    # === Section 2: AFD Marine excerpts ===
    start = 6 + len(passage["waypoints"]) + 2
    afd_marines = forecast.get("afd_marine", {})
    next_row = start
    for office, marine_text in afd_marines.items():
        afd_key = f"afd{office.lower()}"
        afd_label = cycle.get(afd_key, {}).get("label_full", "")
        style_section_header(
            ws.cell(next_row, 1),
            f"AFD{office} Marine excerpt — {afd_label}",
        )
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=5)
        ws.row_dimensions[next_row].height = 22

        afd_cell = ws.cell(next_row + 1, 1, value=marine_text)
        afd_cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        afd_cell.font = Font(name="Calibri", size=10)
        afd_cell.border = thin_border()
        ws.merge_cells(start_row=next_row + 1, start_column=1,
                       end_row=next_row + 1, end_column=5)
        ws.row_dimensions[next_row + 1].height = 130
        next_row += 3

    # === Section 3: Office synopses ===
    for office, syn in forecast.get("synopses", {}).items():
        cwf_key = f"cwf{office.lower()}"
        cwf_label = cycle.get(cwf_key, {}).get("label_short", "")
        style_section_header(
            ws.cell(next_row, 1),
            f"{office} {syn.get('zone', '')} synopsis — {cwf_label}",
        )
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=5)
        ws.row_dimensions[next_row].height = 22

        syn_cell = ws.cell(next_row + 1, 1, value=syn.get("text", ""))
        syn_cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        syn_cell.font = Font(name="Calibri", size=10)
        syn_cell.border = thin_border()
        ws.merge_cells(start_row=next_row + 1, start_column=1,
                       end_row=next_row + 1, end_column=5)
        ws.row_dimensions[next_row + 1].height = 90
        next_row += 3

    # === Section 4: Gulf Stream (optional) ===
    if "gulf_stream" in forecast:
        gs = forecast["gulf_stream"]
        style_section_header(ws.cell(next_row, 1), "Gulf Stream west wall (per NCEP RTOFS)")
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=5)
        ws.row_dimensions[next_row].height = 22

        gs_text = "  |  ".join(gs["positions"])
        gs_cell = ws.cell(next_row + 1, 1,
                          value=f"{gs_text}\n\nEffective {gs['effective_date']}")
        gs_cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        gs_cell.font = Font(name="Calibri", size=10)
        gs_cell.border = thin_border()
        ws.merge_cells(start_row=next_row + 1, start_column=1,
                       end_row=next_row + 1, end_column=5)
        ws.row_dimensions[next_row + 1].height = 70

    set_column_widths(ws, {"A": 8, "B": 26, "C": 32, "D": 12, "E": 60})
    ws.freeze_panes = "A6"


# ==============================================================
# Buoys by WP (static reference)
# ==============================================================
def render_buoys_by_wp(ws, passage):
    """Verification Buoys by Waypoint — which buoys to cross-check at each WP."""
    from ..styles import (
        style_page_title, style_page_subtitle, style_table_header,
        style_centered_cell, style_number_cell, style_text_cell,
        body_bold_font, set_column_widths, fill, thin_border, COLOR_BAND_LIGHT,
    )

    style_page_title(ws.cell(1, 1), "Verification Buoys by Waypoint")
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 26

    style_page_subtitle(
        ws.cell(2, 1),
        "Which NDBC buoy is closest to each waypoint and useful for ground-truth "
        "verification of forecast conditions. Distance is approximate great-circle from "
        "buoy to WP.",
    )
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 32

    headers = ["WP", "Buoy ID", "Buoy Name", "Lat", "Lon", "Distance", "NDBC URL"]
    for i, h in enumerate(headers, 1):
        style_table_header(ws.cell(4, i, value=h))
    ws.row_dimensions[4].height = 28

    r = 5
    band_idx = 0
    for buoy in passage.get("route_buoys", []):
        for wp_id, distance in buoy.get("relevant_wps", {}).items():
            band = COLOR_BAND_LIGHT if band_idx % 2 == 1 else None
            style_centered_cell(ws.cell(r, 1), wp_id, fill_color=band, bold=True)
            style_centered_cell(ws.cell(r, 2), str(buoy["id"]), fill_color=band, bold=True)
            style_text_cell(ws.cell(r, 3), buoy["name"], fill_color=band, wrap=False)
            style_centered_cell(ws.cell(r, 4), f"{buoy['lat']:.3f}°N", fill_color=band)
            style_centered_cell(ws.cell(r, 5), f"{abs(buoy['lon']):.3f}°W", fill_color=band)
            style_text_cell(ws.cell(r, 6), distance, fill_color=band, wrap=False)

            url = f"https://www.ndbc.noaa.gov/station_page.php?station={str(buoy['id']).lower()}"
            url_cell = ws.cell(r, 7, value=url)
            url_cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
            url_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            if band:
                url_cell.fill = fill(band)
            url_cell.border = thin_border()
            ws.row_dimensions[r].height = 22
            r += 1
            band_idx += 1

    set_column_widths(ws, {
        "A": 6, "B": 10, "C": 36, "D": 12, "E": 12, "F": 32, "G": 60,
    })
    ws.freeze_panes = "A5"


# ==============================================================
# Static reference tabs
# ==============================================================
def render_forecast_products(ws, passage):
    """NWS Forecast Products Reference — what each product is and when it's issued."""
    from ..styles import (
        style_page_title, style_page_subtitle, style_table_header,
        style_text_cell, set_column_widths, COLOR_BAND_LIGHT,
    )

    style_page_title(ws.cell(1, 1), "NWS Forecast Products Reference")
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 26

    style_page_subtitle(
        ws.cell(2, 1),
        "What each NWS marine forecast product is and how often it issues. "
        "Use this to decide which source to consult for what question.",
    )
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 28

    headers = ["Product", "Description", "Issuance Cycle", "Issuing Office"]
    for i, h in enumerate(headers, 1):
        style_table_header(ws.cell(4, i, value=h))
    ws.row_dimensions[4].height = 28

    products = [
        ("CWF — Coastal Waters Forecast",
         "Wind, sea, Wave Detail per zone (the workhorse marine product)",
         "4×/day ~4 AM, 10 AM, 4 PM, 10 PM EDT",
         "Office-specific"),
        ("AFD — Area Forecast Discussion",
         "Forecaster narrative + Marine section + Key Messages",
         "4×/day ~3-4 AM, 10-11 AM, 2-3 PM, 8-9 PM",
         "Office-specific"),
        ("OWF — Offshore Waters Forecast",
         "Conditions beyond the 60 NM coastal-zone line",
         "2×/day",
         "Marine Prediction Center"),
        ("High Seas Forecast",
         "GMDSS Area IV — well offshore, blue-water passages",
         "4×/day",
         "Ocean Prediction Center"),
        ("NDBC realtime2",
         "Raw buoy observations, 10-minute updates",
         "Continuous",
         "National Data Buoy Center"),
        ("OPC Surface Analysis",
         "Frontal positions, isobars, sea-state forecast charts",
         "4×/day",
         "Ocean Prediction Center"),
        ("SPC Convective Outlooks",
         "Day 1/2/3 severe weather risk",
         "Multiple daily",
         "Storm Prediction Center"),
        ("NHC Tropical Products",
         "Tropical Weather Outlook, advisories on named storms",
         "Every 6 hr (more often during active systems)",
         "National Hurricane Center"),
    ]
    for i, row in enumerate(products):
        r = 5 + i
        band = COLOR_BAND_LIGHT if i % 2 == 1 else None
        for j, val in enumerate(row, 1):
            style_text_cell(ws.cell(r, j), val, fill_color=band, wrap=True)
        ws.row_dimensions[r].height = 36

    set_column_widths(ws, {"A": 30, "B": 50, "C": 36, "D": 30})
    ws.freeze_panes = "A5"


def render_url_quick_reference(ws, passage=None, buoys=None):
    """Quick-reference URL list, route-driven from passage + buoys YAML."""
    from ..styles import (
        style_page_title, style_page_subtitle, style_section_header,
        style_table_header, style_text_cell, body_bold_font, set_column_widths,
        fill, thin_border, COLOR_BAND_LIGHT,
    )

    style_page_title(ws.cell(1, 1), "Quick Reference — Critical URLs")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 26

    style_page_subtitle(
        ws.cell(2, 1),
        "Bookmark-ready URL list for this passage. Organized by category so you "
        "can jump straight to the right source when a refresh is needed underway.",
    )
    ws.merge_cells("A2:B2")
    ws.row_dimensions[2].height = 28

    # Build URL list organized by section
    sections = []  # [(section_title, [(label, url), ...]), ...]

    # === NWS Offices ===
    office_rows = []
    seen_offices = set()
    if passage and "forecast_offices" in passage:
        for off in passage["forecast_offices"]:
            code = off.get("code", "").upper()
            if not code or code in seen_offices:
                continue
            seen_offices.add(code)
            lower = code.lower()
            cwf_prefix = "fzus51" if code == "AKQ" else "fzus52"
            office_rows.append((f"NWS {code} Marine landing page",
                                f"https://www.weather.gov/{lower}/marine"))
            office_rows.append((f"CWF{code} — raw bulletin text",
                                f"https://tgftp.nws.noaa.gov/data/raw/fz/{cwf_prefix}.k{lower}.cwf.{lower}.txt"))
            office_rows.append((f"AFD{code} — raw discussion text",
                                f"https://tgftp.nws.noaa.gov/data/raw/fx/fxus62.k{lower}.afd.{lower}.txt"))
    if office_rows:
        sections.append(("NWS Forecast Offices", office_rows))

    # === NDBC Buoys ===
    buoy_rows = []
    if buoys and "stations" in buoys:
        for st in buoys["stations"]:
            sid = str(st["id"])
            name = st.get("name", "")
            short_name = name.split("(")[0].strip()
            buoy_rows.append((f"Buoy {sid} — {short_name} (station page)",
                              f"https://www.ndbc.noaa.gov/station_page.php?station={sid.lower()}"))
            buoy_rows.append((f"Buoy {sid} — realtime2 raw feed",
                              f"https://www.ndbc.noaa.gov/data/realtime2/{sid}.txt"))
    if buoy_rows:
        sections.append(("NDBC Buoys", buoy_rows))

    # === Synoptic / Tropical ===
    sections.append(("Synoptic & Tropical", [
        ("OPC Atlantic surface analysis",     "https://ocean.weather.gov/A_sfc_full_ocean.gif"),
        ("OPC Atlantic 24-hr forecast",       "https://ocean.weather.gov/A_24hrwper_full_ocean.gif"),
        ("NHC Atlantic tropical outlook",     "https://www.nhc.noaa.gov/gtwo.php?basin=atlc"),
        ("Saildocs email query (passage)",    "query@saildocs.com"),
    ]))

    # === Tide stations ===
    tide_rows = []
    if passage and "tide_stations" in passage:
        for ts in passage["tide_stations"]:
            label = ts.get("label", f"Tide {ts.get('id', '?')}")
            tid = ts.get("id", "")
            tide_rows.append((label, f"https://tidesandcurrents.noaa.gov/stationhome.html?id={tid}"))
    if tide_rows:
        sections.append(("Tide stations", tide_rows))

    # Render sections
    r = 4
    for section_title, urls in sections:
        style_section_header(ws.cell(r, 1), section_title)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws.row_dimensions[r].height = 24
        r += 1

        # Section column headers
        style_table_header(ws.cell(r, 1, value="Label"))
        style_table_header(ws.cell(r, 2, value="URL"))
        ws.row_dimensions[r].height = 24
        r += 1

        for i, (label, url) in enumerate(urls):
            band = COLOR_BAND_LIGHT if i % 2 == 1 else None
            label_cell = ws.cell(r, 1, value=label)
            label_cell.font = body_bold_font()
            label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            if band:
                label_cell.fill = fill(band)
            label_cell.border = thin_border()

            url_cell = ws.cell(r, 2, value=url)
            url_cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
            url_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            if band:
                url_cell.fill = fill(band)
            url_cell.border = thin_border()
            ws.row_dimensions[r].height = 20
            r += 1
        r += 1  # gap between sections

    set_column_widths(ws, {"A": 44, "B": 90})
    ws.freeze_panes = "A4"


def render_glossary(ws, passage=None):
    """Glossary of terms used in the workbook."""
    from ..styles import (
        style_page_title, style_page_subtitle, style_table_header,
        style_text_cell, body_bold_font, set_column_widths,
        COLOR_BAND_LIGHT,
    )

    vessel_label = (passage or {}).get("vessel", {}).get("designation", "the vessel")

    style_page_title(ws.cell(1, 1), "Glossary")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 26

    style_page_subtitle(
        ws.cell(2, 1),
        "Terms and abbreviations used throughout the workbook. "
        "Boat-specific values reflect the vessel in this passage's YAML.",
    )
    ws.merge_cells("A2:B2")
    ws.row_dimensions[2].height = 24

    # Headers (row 4)
    style_table_header(ws.cell(4, 1, value="Term"))
    style_table_header(ws.cell(4, 2, value="Definition"))
    ws.row_dimensions[4].height = 28

    rows = [
        ("TWA",                "True Wind Angle. Angle between wind direction and boat course, 0-180°."),
        ("TWS",                "True Wind Speed. Wind speed in the earth frame (what a buoy reads)."),
        ("AWA",                "Apparent Wind Angle. What the masthead wind indicator shows; pulled forward by boat motion."),
        ("AWS",                "Apparent Wind Speed. What the masthead anemometer reads."),
        ("BSP",                "Boat Speed through water."),
        ("SOG",                "Speed Over Ground (GPS)."),
        ("CWF",                "Coastal Waters Forecast (NWS marine product, 4× daily)."),
        ("AFD",                "Area Forecast Discussion (NWS narrative, 4× daily)."),
        ("SCA",                "Small Craft Advisory — issued when sustained 18-33 kt, gusts 25-33 kt, or seas ≥6 ft are expected."),
        ("Wave Detail",        "NWS forecast field giving primary swell direction, height, period (and optionally secondary)."),
        ("Sea From",           "Compass direction the primary sea is coming FROM (per NWS Wave Detail)."),
        ("Sea Angle",          "Angle between primary sea direction and boat course (min(|Course-SeaFrom|, 360-|Course-SeaFrom|))."),
        ("Polar Speed",        "Pure boat speed from VPP polar at (TWS, TWA), no sea-state penalty."),
        ("Boat Speed",         "Polar × sea factor (or motor speed if below crossover). This drives ETAs."),
        ("Sea Factor",         "Multiplier applied to polar to account for sea-state degradation. Project values: 0.72 / 0.87 / 0.92."),
        ("Motor crossover TWS", f"Wind speed below which motor is faster than sailing. {vessel_label}: ~7 kt."),
        ("Code D",             f"Cruising downwind sail. {vessel_label} usable range AWS 6-18 kt, TWA 90-150°."),
        ("Civil dawn / dusk",  "When sun is 6° below horizon. Defines twilight transitions for arrival color coding."),
        ("Hs",                 "Significant wave height (average of highest 1/3 of waves)."),
        ("WMO header",         "Standard weather product timestamp — e.g. FZUS52 KCHS 121518 = day 12 @ 1518 UTC."),
        ("ETA color key",      "Green = full daylight. Yellow = civil twilight. Red = night arrival."),
        ("Row band",           "Whole-row tint indicating non-color-coded operational concerns (tactical hazards, etc.)."),
    ]
    for i, (term, defn) in enumerate(rows):
        r = 5 + i
        # Alternating bands for visual rhythm
        band_color = COLOR_BAND_LIGHT if i % 2 == 1 else None
        term_cell = ws.cell(r, 1, value=term)
        term_cell.font = body_bold_font()
        term_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        from ..styles import fill, thin_border
        if band_color:
            term_cell.fill = fill(band_color)
        term_cell.border = thin_border()

        style_text_cell(ws.cell(r, 2), defn, fill_color=band_color)
        ws.row_dimensions[r].height = 22

    set_column_widths(ws, {"A": 22, "B": 100})
    ws.freeze_panes = "A5"


# ==============================================================
# Format Reference
# ==============================================================
def _render_freshness_panel_legacy(ws, forecast, buoys):
    """Legacy freshness panel — reads from forecast['cycle'] and buoys['stations'].

    Used when build is called without a weather_pull status.json. Maintained for
    backward compatibility; new builds should pass status_data to get the
    authoritative view of source provenance.
    """
    ws.cell(3, 1, value="⚠ DATA FRESHNESS PANEL — refresh before any GO/NO-GO decision").font = Font(bold=True, color=COLOR_DANGER_FONT)
    headers = ["Source ID", "Kind", "Issuance / Obs Time", "Age", "Status", "Provenance"]
    for i, h in enumerate(headers, 1):
        style_header_cell(ws.cell(4, i, value=h))

    cycle = forecast["cycle"]
    now_utc = datetime.now(timezone.utc)
    freshness_rows = []
    for key, val in cycle.items():
        if key.startswith("cwf") and isinstance(val, dict):
            office = key[3:].upper()
            age_str, status, color = classify_freshness(val.get("issued"), now_utc)
            freshness_rows.append(
                (f"cwf:{office}", "CWF", val.get("label_full", ""),
                 age_str, status, color,
                 f"tgftp.nws.noaa.gov/.../{office.lower()}.txt")
            )
    for key, val in cycle.items():
        if key.startswith("afd") and isinstance(val, dict):
            office = key[3:].upper()
            age_str, status, color = classify_freshness(val.get("issued"), now_utc)
            freshness_rows.append(
                (f"afd:{office}", "AFD", val.get("label_full", ""),
                 age_str, status, color,
                 f"forecast.weather.gov/...AFD{office}")
            )
    for i, frow in enumerate(freshness_rows):
        source, product, issuance, age_str, status, color, url = frow
        cells = [source, product, issuance, age_str, status, url]
        for j, val in enumerate(cells, 1):
            cell = ws.cell(5 + i, j, value=val)
            if j == 5 and color:
                cell.fill = fill(color)
                cell.font = Font(bold=True)

    # Buoy rows
    buoy_row = 5 + len(freshness_rows) + 1
    for st in buoys["stations"]:
        if st.get("status") == "offline":
            row = (f"buoy:{st['id']} ({st['name']})", "BUOY",
                   f"OFFLINE since {st.get('offline_since', '')}", "—", "OFFLINE", "—")
        else:
            status_text = st.get("status", "fresh")
            row = (f"buoy:{st['id']} ({st['name']})", "BUOY",
                   st.get("reading_time", ""), "<2 hr",
                   status_text.replace("_", " ").upper(),
                   f"ndbc.noaa.gov/data/realtime2/{st['id']}.txt")
        for j, val in enumerate(row, 1):
            ws.cell(buoy_row, j, value=val)
        buoy_row += 1
    return buoy_row


def _render_freshness_panel_from_status(ws, status_data):
    """Authoritative freshness panel — reads from weather_pull _status.json.

    Every source in the manifest gets a row. Provenance shows the URL that
    succeeded, "(user paste)" for ingested pastes, or "NOT COVERED" with red
    fill for required-missing sources. Age and tier come directly from the
    weather_pull assessment, so what you see is exactly what was used.
    """
    pulled_at = status_data.get("pulled_at", "")
    pulled_str = pulled_at[:16].replace("T", " ") + " UTC" if pulled_at else ""
    title = "⚠ DATA FRESHNESS PANEL — built from weather_pull status"
    if pulled_str:
        title += f"  (pulled {pulled_str})"
    ws.cell(3, 1, value=title).font = Font(bold=True, color=COLOR_DANGER_FONT)
    headers = ["Source ID", "Kind", "Issuance / Obs Time", "Age", "Status", "Provenance"]
    for i, h in enumerate(headers, 1):
        style_header_cell(ws.cell(4, i, value=h))

    row = 5
    for src in status_data.get("sources", []):
        sid = src["id"]
        kind = src["kind"].upper()
        location = src.get("location", "")
        # Friendly identifier — include location for buoys, just office for bulletins
        if kind == "BUOY" and location:
            source_label = f"{sid} ({location.split('(')[0].strip()})"
        else:
            source_label = sid

        if src.get("covered"):
            tier = src.get("tier", "UNKNOWN")
            age = src.get("age_str", "?")
            issued = src.get("issued_utc", "")
            if issued:
                # Trim ISO to readable: 2026-05-12T18:30+00:00 -> 2026-05-12 18:30 UTC
                issued_str = issued[:16].replace("T", " ") + " UTC"
            else:
                issued_str = "(parsed, no timestamp)"
            url = src.get("accepted_url", "")
            if url == "user_paste":
                provenance = "(user paste)"
            elif url == "drive_sync":
                provenance = "(Drive sync)"
            elif url:
                provenance = url.replace("https://", "").replace("http://", "")
                if len(provenance) > 55:
                    provenance = provenance[:52] + "..."
            else:
                provenance = ""
            color = _tier_color(tier)
        else:
            tier = "MISSING" if src.get("required") else "SKIPPED"
            age = "—"
            issued_str = "NOT COVERED"
            n_attempts = len(src.get("attempts", []))
            provenance = f"tried {n_attempts} URL{'s' if n_attempts != 1 else ''}, no fresh data"
            color = COLOR_BAD_FILL if src.get("required") else None

        cells = [source_label, kind, issued_str, age, tier, provenance]
        for j, val in enumerate(cells, 1):
            cell = ws.cell(row, j, value=val)
            if j == 5 and color:
                cell.fill = fill(color)
                cell.font = Font(bold=True)
        row += 1
    return row + 1  # leave a blank row below the panel


def _tier_color(tier):
    """Map weather_pull tier vocab to workbook fill colors."""
    return {
        "FRESH": COLOR_GOOD_FILL,
        "STALE": COLOR_NEUTRAL_FILL,
        "ARCHIVED": COLOR_BAD_FILL,
        "UNKNOWN": COLOR_NEUTRAL_FILL,
        "MISSING": COLOR_BAD_FILL,
    }.get(tier)


# Format Reference
# ==============================================================
def render_format_reference(ws, passage, forecast, buoys, legs_by_plan, status_data=None):
    """Render the Format Reference tab.

    The freshness panel reads from one of two sources:

    1. If `status_data` is provided (from weather_pull's _status.json),
       it is treated as authoritative: every source in the manifest
       gets a row, including those that failed to cover. Provenance
       shows the URL that succeeded, or "(user paste)" for ingested
       pastes, or "NOT COVERED" for required-missing sources.

    2. Otherwise (legacy path), the panel is derived from the
       `forecast['cycle']` block and `buoys['stations']` list. This is
       backward-compatible with workbooks built without --require.
    """
    ws.cell(1, 1, value="Format Reference — workbook conventions, calibration, color keys").font = Font(size=14, bold=True)
    ws.merge_cells("A1:F1")

    # Data freshness panel
    if status_data is not None:
        buoy_row = _render_freshness_panel_from_status(ws, status_data)
    else:
        buoy_row = _render_freshness_panel_legacy(ws, forecast, buoys)

    # Sea factor calibration
    sf = passage["calibration"]["sea_factors"]
    sf_row = buoy_row + 2
    ws.cell(sf_row, 1, value=f"{passage['vessel']['designation']} Calibration").font = Font(bold=True)
    sf_block = [
        ("Motor speed", f"{passage['vessel']['motor_speed_kt']} kt"),
        ("Motor crossover TWS", f"{passage['vessel']['motor_crossover_tws_kt']} kt"),
        ("Sea factor — steep chop on bow (TWA<60°, period<6s, Hs/T≥0.8)", f"{sf['steep_chop_bow']}"),
        ("Sea factor — close reach (TWA 60-80°)", f"{sf['close_reach']}"),
        ("Sea factor — beam reach (TWA 80-100°)", f"{sf['beam_reach']}"),
        ("Sea factor — broad reach (TWA 100-150°)", f"{sf['broad_reach']}"),
        ("Sea factor — reefed in active trough", f"{sf['reefed_active_trough']}"),
    ]
    for i, (param, val) in enumerate(sf_block):
        ws.cell(sf_row + 1 + i, 1, value=param)
        ws.cell(sf_row + 1 + i, 2, value=val)

    # ETA color key
    eta_row = sf_row + len(sf_block) + 3
    ws.cell(eta_row, 1, value="ETA Color Key").font = Font(bold=True)
    _legend_block(ws, eta_row + 1, COLOR_GOOD_FILL, "DAY", "6:35 AM - 8:10 PM (sun well above horizon)")
    _legend_block(ws, eta_row + 2, COLOR_NEUTRAL_FILL, "TWILIGHT", "5:55-6:35 AM, 8:10-8:45 PM (civil twilight)")
    _legend_block(ws, eta_row + 3, COLOR_BAD_FILL, "NIGHT", "8:45 PM - 5:55 AM (dark)")

    # Reverse arrival calculator
    rc_row = eta_row + 5
    ws.cell(rc_row, 1, value="Reverse Arrival Calculator").font = Font(bold=True)
    rc_band_blue = "DDEBF7"
    rc_band_gray = "F2F2F2"
    # Constant inputs (blue)
    ws.cell(rc_row + 1, 1, value="Total route distance (NM)").fill = fill(rc_band_blue)
    ws.cell(rc_row + 1, 2, value=passage["passage"]["total_nm"]).fill = fill(rc_band_blue)
    ws.cell(rc_row + 1, 3, value="Target arrival window").fill = fill(rc_band_blue)
    at = passage["arrival_timing"]
    ws.cell(rc_row + 1, 4, value=f"Civil dawn {_h(at['civil_dawn'])} → sunrise {_h(at['sunrise'])} → preferred {_h(at['preferred_window']['start'])}-{_h(at['preferred_window']['end'])}").fill = fill(rc_band_blue)

    # Plan rows — derived values (gray)
    for plan_idx, plan in enumerate(passage["plans"]):
        legs = legs_by_plan.get(plan["id"], [])
        if not legs:
            continue
        arrival_leg = legs[-1]
        avg_speed = passage["passage"]["total_nm"] / arrival_leg.cum_time_hr if arrival_leg.cum_time_hr else 0
        r = rc_row + 2 + plan_idx
        for c in [1, 2, 3, 4]:
            ws.cell(r, c).fill = fill(rc_band_gray)
        ws.cell(r, 1, value=f"{plan['tab_label']} (avg {avg_speed:.1f} kt)")
        ws.cell(r, 3, value=f"Depart {plan['depart_day']} {_h(plan['depart_hour'])} → Arrival {arrival_leg.eta_str} ({arrival_leg.cum_time_hr:.1f} hr)")
        color_text = {"C6EFCE": "DAY ✓", "FFEB9C": "TWILIGHT/DAWN ✓", "FFC7CE": "⚠ NIGHT"}.get(arrival_leg.eta_color, "")
        ws.cell(r, 4, value=color_text)
        # Also color the arrival window cell by ETA color
        ws.cell(r, 4).fill = fill(arrival_leg.eta_color)

    # Final column widths — sized to fit ALL sections of this tab without overflow:
    #   - freshness panel: A=source ID (~32), B=kind (~10), C=issuance (~22),
    #                      D=age (~10), E=status (~12), F=provenance URL (~60)
    #   - sea factor calibration: A=description (~60 for longest), B=value
    #   - ETA color key: A=chip ("DAY"), B=description (~45)
    #   - reverse calculator: A-D used for arrival timing labels/values
    for col, w in [("A", 50), ("B", 46), ("C", 26), ("D", 50), ("E", 14), ("F", 60)]:
        ws.column_dimensions[col].width = w


def _legend_block(ws, row, color, label, desc):
    chip = ws.cell(row, 1, value=label)
    chip.fill = fill(color)
    chip.font = Font(bold=True)
    chip.alignment = Alignment(horizontal="center")
    ws.cell(row, 2, value=desc)


def _h(decimal_h):
    """Format decimal hour as 'H:MM AM/PM'."""
    h = int(decimal_h)
    m = int(round((decimal_h - h) * 60))
    if m == 60:
        h += 1
        m = 0
    if h == 0:
        return f"12:{m:02d} AM"
    if h < 12:
        return f"{h}:{m:02d} AM"
    if h == 12:
        return f"12:{m:02d} PM"
    return f"{h-12}:{m:02d} PM"


# ==============================================================
# Refresh Cadence (template)
# ==============================================================
def render_refresh_cadence(ws, passage=None):
    """Forecast refresh cadence reference — when each product issues."""
    from ..styles import (
        style_page_title, style_page_subtitle, style_section_header,
        style_table_header, style_text_cell, style_centered_cell,
        body_bold_font, set_column_widths, fill, thin_border,
        COLOR_BAND_LIGHT,
    )

    style_page_title(ws.cell(1, 1), "Forecast Refresh Cadence")
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 26

    style_page_subtitle(
        ws.cell(2, 1),
        "NWS forecast bulletins issue 4×/day; pulled data goes stale within 6–12 hours. "
        "Data hierarchy by freshness: (1) NDBC realtime2 buoys <1 hr — authoritative for current state. "
        "(2) AFD current cycle — narrative + key messages, 4×/day. (3) CWF current cycle — wind/sea/Wave Detail, 4×/day.",
    )
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 48

    # === Section 1: NWS issuance schedule ===
    style_section_header(ws.cell(4, 1), "NWS standing issuance schedule (EDT, approximate)")
    ws.merge_cells("A4:F4")
    ws.row_dimensions[4].height = 22

    headers = ["Product", "Office", "Cycle 1", "Cycle 2", "Cycle 3", "Cycle 4"]
    for i, h in enumerate(headers, 1):
        style_table_header(ws.cell(5, i, value=h))
    ws.row_dimensions[5].height = 28

    schedule = []
    seen_offices = []
    if passage and "forecast_offices" in passage:
        for off in passage["forecast_offices"]:
            code = off.get("code", "")
            office_name = off.get("office", "")
            if code and code not in seen_offices:
                schedule.append((f"CWF{code}", office_name, "~4 AM", "~10 AM", "~4 PM", "~10 PM"))
                seen_offices.append(code)
    if not schedule:
        schedule = [
            ("CWFCHS", "NWS Charleston SC", "~4 AM", "~10 AM", "~4 PM", "~10 PM"),
        ]
    # Always add OPC and NDBC
    schedule.append(("OPC offshore", "Ocean Prediction Center", "~12 AM", "~6 AM", "~12 PM", "~6 PM"))
    schedule.append(("NDBC buoys",   "National Data Buoy Center", "Continuous", "10–60 min", "—", "—"))

    for i, row_data in enumerate(schedule):
        r = 6 + i
        band = COLOR_BAND_LIGHT if i % 2 == 1 else None
        product, office, c1, c2, c3, c4 = row_data
        prod_cell = ws.cell(r, 1, value=product)
        prod_cell.font = body_bold_font()
        prod_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        if band:
            prod_cell.fill = fill(band)
        prod_cell.border = thin_border()
        style_text_cell(ws.cell(r, 2), office, fill_color=band, wrap=False)
        for j, val in enumerate([c1, c2, c3, c4], 3):
            style_centered_cell(ws.cell(r, j), val, fill_color=band)
        ws.row_dimensions[r].height = 22

    # === Section 2: Standing workflow ===
    wf_row = 6 + len(schedule) + 2
    style_section_header(ws.cell(wf_row, 1), "Standing refresh workflow")
    ws.merge_cells(start_row=wf_row, start_column=1, end_row=wf_row, end_column=6)
    ws.row_dimensions[wf_row].height = 22

    workflow = (
        "AT EACH REFRESH:\n"
        "1. Pull current NDBC realtime2 data for relevant buoys.\n"
        "2. Pull current AFD (each office along the route).\n"
        "3. Pull current CWF for each zone the route crosses.\n"
        "4. Update Live Buoy Data, Forecast Sources, and Plan tabs.\n"
        "5. Re-evaluate GO/NO-GO decision against new data.\n"
        "6. Add Verification Scorecard entry if material change detected."
    )
    wf_cell = ws.cell(wf_row + 1, 1, value=workflow)
    wf_cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
    wf_cell.font = Font(name="Calibri", size=10)
    wf_cell.border = thin_border()
    ws.merge_cells(start_row=wf_row + 1, start_column=1, end_row=wf_row + 1, end_column=6)
    ws.row_dimensions[wf_row + 1].height = 130

    set_column_widths(ws, {"A": 22, "B": 38, "C": 14, "D": 14, "E": 14, "F": 14})
    ws.freeze_panes = "A6"


# ==============================================================
# Verification Scorecard (template + lessons)
# ==============================================================
def render_verification_scorecard(ws, passage, lessons):
    """Verification Scorecard — methodology lessons + forecast-skill tracking template."""
    from ..styles import (
        style_page_title, style_page_subtitle, style_section_header,
        style_table_header, body_bold_font, set_column_widths,
        fill, thin_border, COLOR_BAND_LIGHT, COLOR_TITLE_FONT,
    )

    style_page_title(ws.cell(1, 1), f"Verification Scorecard — {passage['passage']['name']}")
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 26

    style_page_subtitle(
        ws.cell(2, 1),
        "Methodology lessons and forecast-skill tracking. Lessons codify what "
        "we've learned to improve future passages. Skill tracking compares forecast "
        "calls against observed buoy data and underway logs.",
    )
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 32

    # === Section: Methodology Lessons ===
    style_section_header(ws.cell(4, 1), "Methodology Lessons")
    ws.merge_cells("A4:G4")
    ws.row_dimensions[4].height = 22

    # Render each lesson as a 2-row block: headline row + body row
    cur = 5
    for i, lesson in enumerate(lessons):
        band = COLOR_BAND_LIGHT if i % 2 == 1 else None

        # Headline row
        id_cell = ws.cell(cur, 1, value=lesson["id"])
        id_cell.font = Font(name="Calibri", size=11, bold=True, color=COLOR_TITLE_FONT)
        id_cell.alignment = Alignment(horizontal="center", vertical="center")
        id_cell.border = thin_border()
        if band:
            id_cell.fill = fill(band)

        head_cell = ws.cell(cur, 2, value=lesson["headline"])
        head_cell.font = Font(name="Calibri", size=11, bold=True, color=COLOR_TITLE_FONT)
        head_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        head_cell.border = thin_border()
        if band:
            head_cell.fill = fill(band)
        ws.merge_cells(start_row=cur, start_column=2, end_row=cur, end_column=7)
        ws.row_dimensions[cur].height = 22

        # Body row
        body_cell = ws.cell(cur + 1, 2, value=lesson["body"])
        body_cell.font = Font(name="Calibri", size=10)
        body_cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        body_cell.border = thin_border()
        if band:
            body_cell.fill = fill(band)
        # Empty styled cell in col 1 for visual continuity
        gutter = ws.cell(cur + 1, 1, value="")
        gutter.border = thin_border()
        if band:
            gutter.fill = fill(band)
        ws.merge_cells(start_row=cur + 1, start_column=2, end_row=cur + 1, end_column=7)
        ws.row_dimensions[cur + 1].height = max(50, min(180, len(lesson["body"]) // 4))

        cur += 3  # 2 rows + 1 spacer

    set_column_widths(ws, {
        "A": 12, "B": 28, "C": 22, "D": 22, "E": 22, "F": 22, "G": 22,
    })
    ws.freeze_panes = "A5"


# ==============================================================
# Vessel Comparison — HR 48 Mk II vs HR 54 (Frers VPP polars, GZ stability)
# Mandatory in every workbook. Five sections:
#   1. Side-by-side polar grid at key TWA × TWS
#   2. Static comfort & safety metrics from data sheets
#   3. Stability (GZ curves, RM @ 30°, AVS)
#   4. Per-leg HR 48 vs HR 54 polar speed projection for THIS passage
#   5. Practical takeaways for this passage
# ==============================================================
def render_vessel_comparison(ws, passage, legs_by_plan):
    """Side-by-side HR 48 vs HR 54 comparison.

    `legs_by_plan` is {plan_id: list[Leg]} — used to drive Section 4 per-leg
    speed projection. Falls back to Section 4 being empty if not provided.
    """
    from ..polar import polar_speed

    ws.cell(1, 1, value="Vessel Comparison — HR 48 Mk II vs HR 54").font = Font(size=14, bold=True)
    ws.merge_cells("A1:I1")
    ws.cell(2, 1, value=(
        "Sources: /mnt/project/HR48Boatspeeds.xls (D1170 half-load), "
        "/mnt/project/HR54speedtable.xls (D1206 half-load), "
        "/mnt/project/HR48_Stability.pdf (Light Ship), "
        "/mnt/project/54_Stability.pdf (Full Load), "
        "/mnt/project/Data_sheet_40Mk2_48Mk2_55.pdf, "
        "/mnt/project/54_Data_Sheet.pdf"
    )).font = Font(size=9, italic=True)
    ws.merge_cells("A2:I2")

    # === Section 1: Polar grid HR48 vs HR54 at key TWA × TWS ===
    r = 4
    ws.cell(r, 1, value="Section 1 — Polar speed (kt) at key TWA × TWS, Frers VPP half-load").font = Font(bold=True)
    r += 1
    headers = ["TWA °", "TWS 6  HR48", "TWS 6  HR54", "TWS 10 HR48", "TWS 10 HR54",
               "TWS 14 HR48", "TWS 14 HR54", "TWS 20 HR48", "TWS 20 HR54"]
    for i, h in enumerate(headers, 1):
        style_header_cell(ws.cell(r, i, value=h))
    r += 1

    twa_grid = [45, 52, 60, 70, 80, 90, 100, 110, 120, 135, 150]
    tws_pairs = [6, 10, 14, 20]
    for twa in twa_grid:
        ws.cell(r, 1, value=float(twa)).number_format = "0"
        col = 2
        for tws in tws_pairs:
            v48 = polar_speed(tws, twa, "D1170")
            v54 = polar_speed(tws, twa, "D1206")
            ws.cell(r, col, value=round(v48, 2)).number_format = "0.00"
            ws.cell(r, col + 1, value=round(v54, 2)).number_format = "0.00"
            col += 2
        r += 1

    # === Section 2: Static comfort & safety metrics ===
    r += 2
    ws.cell(r, 1, value="Section 2 — Static comfort & safety metrics (from data sheets)").font = Font(bold=True)
    r += 1
    for i, h in enumerate(["Metric", "HR 48 Mk II", "HR 54", "Scale / Interpretation"], 1):
        style_header_cell(ws.cell(r, i, value=h))
    r += 1
    static_metrics = [
        ("LOA",                       "47.7 ft (14.55 m)",     "53.0 ft (16.16 m)",   "Length overall"),
        ("LWL",                       "43.5 ft (13.25 m)",     "47.0 ft (14.34 m)",   "Waterline"),
        ("Beam",                      "14.6 ft (4.45 m)",      "15.3 ft (4.65 m)",    "Max beam"),
        ("Draft (deep keel)",         "7.7 ft (2.35 m)",       "8.2 ft (2.50 m)",     ""),
        ("Air draft",                 "71.4 ft (21.77 m)",     "80.4 ft (24.50 m)",   "Mast height incl. antennas"),
        ("Empty displacement",        "40,785 lb (18.5 t)",    "47,400 lb (21.5 t)",  "Per data sheet"),
        ("Sailing displacement (half-load)", "~45,200 lb (20.5 t)", "~49,600 lb (22.5 t)", "Empty + 0.5 × max load"),
        ("Ballast ratio",             "41.9%",                  "37.1%",              "HR 48 stiffer per pound"),
        ("Brewer Comfort Ratio",      "42.9",                   "52.1",               ">50 = heavy cruiser, 40-50 = moderate"),
        ("Capsize Screening Formula", "1.66",                   "1.55",               "<2.0 acceptable offshore; lower safer"),
        ("D/L ratio (half-load)",     "246",                    "274",                "Both moderate cruisers (200-300)"),
        ("SA/D ratio (half-load)",    "18.9",                   "19.0",               "Power-to-weight virtually identical"),
        ("Hull speed",                "8.83 kt",                "9.18 kt",            "1.34 × √LWL"),
        ("Upwind sail area",          "144.7 m² (1558 sq ft)",  "176.6 m² (1901 sq ft)", "Main + genoa"),
        ("Engine",                    "Volvo D3-110",           "Volvo D4-180",       ""),
        ("Fuel capacity",             "110 gal (420 L)",        "238 gal (900 L)",    ""),
        ("Motor range at 6 kt",       "~440 NM",                "~830 NM",            "HR 54 carries 2× fuel"),
    ]
    for metric, v48, v54, note in static_metrics:
        ws.cell(r, 1, value=metric)
        ws.cell(r, 2, value=v48)
        ws.cell(r, 3, value=v54)
        ws.cell(r, 4, value=note)
        r += 1

    # === Section 3: Stability (GZ curves) ===
    r += 1
    ws.cell(r, 1, value="Section 3 — Stability (GZ curves from Frers/Hallberg-Rassy)").font = Font(bold=True)
    r += 1
    for i, h in enumerate(["Metric", "HR 48 Mk II", "HR 54", "Note"], 1):
        style_header_cell(ws.cell(r, i, value=h))
    r += 1
    stability_metrics = [
        ("RMC (righting moment coeff.)", "467 kg·m",     "656 kg·m",          "HR 48 from Light Ship sheet, HR 54 from Full Load sheet"),
        ("RM @ 30°",                     "11,995 kg·m",  "~16,500 kg·m est.", "HR 54 PDF reports RM @ 25° = 14,619 kg·m; +13% @ 30° typical"),
        ("RM max",                       "17,657 kg·m",  "23,654 kg·m",       "Both peak around 70° heel"),
        ("AVS (vanishing stability)",    "≈125°",        "≈125-128°",         "Both well above ISO Cat A 120° threshold"),
        ("RM advantage at working range","—",            "+14-20%",           "Apples-to-apples: HR 54 has 14-20% more righting moment than HR 48 in 20-40° heel"),
        ("Load state of source PDF",     "Light Ship (HR48_Stability.pdf)", "Full Load (54_Stability.pdf)", "Caveat: Light Ship typically -7-10% RM vs Full Load"),
    ]
    for metric, v48, v54, note in stability_metrics:
        ws.cell(r, 1, value=metric)
        ws.cell(r, 2, value=v48)
        ws.cell(r, 3, value=v54)
        ws.cell(r, 4, value=note)
        r += 1

    # === Section 4: Per-leg HR48 vs HR54 polar speed projection ===
    r += 1
    ws.cell(r, 1, value="Section 4 — Per-leg HR 48 vs HR 54 polar speeds (this passage)").font = Font(bold=True)
    r += 1
    ws.cell(r, 1, value="Pure polar speed (no sea factor). Boat Speed in Plan tabs applies sea factor.").font = Font(size=9, italic=True)
    r += 1
    for i, h in enumerate(["WP", "TWA", "TWS", "HR 48 polar", "HR 54 polar", "Δ kt", "Δ %"], 1):
        style_header_cell(ws.cell(r, i, value=h))
    r += 1

    # Use first plan's legs as the projection basis
    first_plan_legs = None
    if legs_by_plan:
        first_plan_id = next(iter(legs_by_plan))
        first_plan_legs = legs_by_plan[first_plan_id]

    if first_plan_legs:
        total_d48 = 0.0
        total_d54 = 0.0
        n = 0
        for idx, leg in enumerate(first_plan_legs):
            if leg.course_out is None:
                continue  # terminal WP has no leg metrics
            next_wp_id = first_plan_legs[idx + 1].wp_id if idx + 1 < len(first_plan_legs) else None
            twa = leg.twa
            tws = (leg.wind_kt_low + leg.wind_kt_high) / 2.0
            v48 = polar_speed(tws, twa, "D1170")
            v54 = polar_speed(tws, twa, "D1206")
            delta_kt = v54 - v48
            delta_pct = (delta_kt / v48 * 100) if v48 > 0 else 0.0
            ws.cell(r, 1, value=f"{leg.wp_id}→{next_wp_id}" if next_wp_id else leg.wp_id)
            ws.cell(r, 2, value=float(twa)).number_format = "0"
            ws.cell(r, 3, value=round(tws, 1)).number_format = "0.0"
            ws.cell(r, 4, value=round(v48, 2)).number_format = "0.00"
            ws.cell(r, 5, value=round(v54, 2)).number_format = "0.00"
            ws.cell(r, 6, value=round(delta_kt, 2)).number_format = "0.00"
            ws.cell(r, 7, value=f"{'+' if delta_pct >= 0 else ''}{delta_pct:.1f}%")
            total_d48 += v48
            total_d54 += v54
            n += 1
            r += 1
        # AVG row
        if n > 0:
            avg48 = total_d48 / n
            avg54 = total_d54 / n
            avg_delta = avg54 - avg48
            avg_pct = (avg_delta / avg48 * 100) if avg48 > 0 else 0.0
            ws.cell(r, 1, value="AVG").font = Font(bold=True)
            ws.cell(r, 4, value=round(avg48, 2)).number_format = "0.00"
            ws.cell(r, 4).font = Font(bold=True)
            ws.cell(r, 5, value=round(avg54, 2)).number_format = "0.00"
            ws.cell(r, 5).font = Font(bold=True)
            ws.cell(r, 6, value=round(avg_delta, 2)).number_format = "0.00"
            ws.cell(r, 6).font = Font(bold=True)
            ws.cell(r, 7, value=f"{'+' if avg_pct >= 0 else ''}{avg_pct:.1f}%").font = Font(bold=True)
            r += 1

    # === Section 5: Practical takeaways ===
    r += 1
    ws.cell(r, 1, value="Section 5 — Practical takeaways for this passage").font = Font(bold=True)
    r += 1

    takeaways = []
    if first_plan_legs and n > 0:
        savings_pct = avg_pct
        # Estimate time savings: passage time * (1 - 1/(1+pct/100))
        last_leg = first_plan_legs[-1]
        passage_hr = last_leg.cum_time_hr if hasattr(last_leg, 'cum_time_hr') else 0
        if passage_hr > 0 and savings_pct > 0:
            time_savings_hr = passage_hr * (savings_pct / (100 + savings_pct))
            takeaways.append(
                f"• Average polar advantage HR 54 over HR 48 across this passage's wind/angle mix: +{savings_pct:.1f}%"
            )
            takeaways.append(
                f"• If this passage were sailed in an HR 54, total time would drop by approximately "
                f"{time_savings_hr:.1f} hr ({passage_hr:.1f} hr → {passage_hr - time_savings_hr:.1f} hr) at the polar level."
            )
    takeaways.extend([
        "• HR 54 stability margin: +14-20% more righting moment than HR 48 in the working 20-40° heel range. Translates to higher reefing thresholds in heavy weather.",
        "• HR 48 fuel range ~440 NM vs HR 54 ~830 NM. HR 48 must motor carefully through light spells; HR 54 has substantial range reserve for contingencies.",
        "• HR 48 ballast ratio 41.9% > HR 54 37.1% — HR 48 is relatively stiffer per pound of displacement, but absolute righting moment still favors HR 54.",
        "• Biggest HR 54 advantage typically shows on broad-reach and beam-reach legs in 10-20 kt TWS where the larger sail plan and waterline length compound.",
    ])
    for line in takeaways:
        cell = ws.cell(r, 1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        ws.row_dimensions[r].height = 30
        r += 1

    # Column widths
    for col, w in [("A", 28), ("B", 22), ("C", 22), ("D", 32), ("E", 14), ("F", 14), ("G", 14), ("H", 14), ("I", 14)]:
        ws.column_dimensions[col].width = w
