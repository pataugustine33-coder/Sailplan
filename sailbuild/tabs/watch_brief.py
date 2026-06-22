"""Watch Brief tab — 12-hour tactical dashboard, one per plan.

Layout (Option C from skipper approval Mon 6/22):
  - Header (rows 1-5): plan label, time window, position, ETA, cycle source
  - Watch cards image (rows 7-30): 4 segments side-by-side, single PNG
  - 12-hour strip image (rows 32-50): wind/sea/boat-speed across the 12 hr
  - Tactical actions table (rows 52+): one row per segment with action text

Designed for at-sea watch-handover briefings. Boat speed is shown as BOTH
polar potential (clean polar Vs) AND calibrated expected (after sea_factor),
so the skipper sees the upper bound and the realistic expected speed.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from ..compute import build_watch_segments
from ..charts import watch_cards_png_bytes, twelve_hour_strip_png_bytes


THIN = Side(border_style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="1F3864")
SUBHEADER_FONT = Font(name="Calibri", size=10, color="333333")

ACTION_LABEL_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
ACTION_TEXT_FONT = Font(name="Calibri", size=10, color="333333")

RISK_FILL = {
    "green":  PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "yellow": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "red":    PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}
RISK_LABEL_FILL = {
    "green":  PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
    "yellow": PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid"),
    "red":    PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
}


def render_watch_brief_tab(ws, passage, plan_meta, legs, total_nm, forecast_cycle_label=""):
    """Build the Watch Brief tab content into the given worksheet.

    Args:
      ws: the openpyxl worksheet (already created and sized)
      passage: full passage YAML dict (for arrival_timing, vessel info)
      plan_meta: plan block from passage YAML (depart_day, depart_hour, etc.)
      legs: list of Leg objects from build_legs_for_plan
      total_nm: total passage distance NM
      forecast_cycle_label: short label of the forecast cycle (for header)
    """
    arrival_timing = passage.get("arrival_timing")
    segments = build_watch_segments(plan_meta, legs, arrival_timing, total_nm)

    # Column widths — let cells span the image widths reasonably
    ws.column_dimensions["A"].width = 4
    for col in range(2, 18):
        ws.column_dimensions[get_column_letter(col)].width = 11

    # ====================================================================
    # SECTION 1 — HEADER (rows 1-5)
    # ====================================================================
    plan_label = plan_meta.get("tab_label", "Plan A")
    vessel = passage["vessel"]["designation"]
    title = f"WATCH BRIEF — {vessel} — {plan_label}"
    ws.cell(row=1, column=2, value=title).font = TITLE_FONT
    ws.cell(row=1, column=2).fill = TITLE_FILL
    ws.cell(row=1, column=2).alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=17)
    ws.row_dimensions[1].height = 26

    # Subtitle: time window
    if segments:
        window_label = f"Next 12 hours: {segments[0].start_clock} → {segments[-1].end_clock} EDT"
    else:
        window_label = "Next 12 hours"
    ws.cell(row=2, column=2, value=window_label).font = Font(
        name="Calibri", size=11, italic=True, color="1F3864")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=17)

    # Context line: position, distance, ETA at polar, cycle
    eta_label = legs[-1].eta_str if legs else "—"
    origin = passage["passage"].get("origin", "")
    destination = passage["passage"].get("destination", "")
    ctx_parts = [
        f"From: {origin[:50]}",
        f"To: {destination[:40]}",
        f"Polar ETA: {eta_label}",
    ]
    if forecast_cycle_label:
        ctx_parts.append(f"Cycle: {forecast_cycle_label}")
    ws.cell(row=3, column=2, value="  ·  ".join(ctx_parts)).font = SUBHEADER_FONT
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=17)

    # Distance remaining + total
    ws.cell(row=4, column=2,
            value=f"Total passage: {total_nm:.1f} NM"
            ).font = SUBHEADER_FONT
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=17)

    # ====================================================================
    # SECTION 2 — 4 WATCH CARDS (single PNG image)
    # ====================================================================
    cards_buf = watch_cards_png_bytes(segments, output_w_px=1600, output_h_px=540)
    if cards_buf:
        img = XLImage(cards_buf)
        img.width = 1200
        img.height = 405
        img.anchor = "B6"
        ws.add_image(img)

    # ====================================================================
    # SECTION 3 — 12-HOUR STRIP (single PNG image)
    # ====================================================================
    strip_buf = twelve_hour_strip_png_bytes(segments, output_w_px=1400, output_h_px=380)
    strip_anchor_row = 28
    if strip_buf:
        img = XLImage(strip_buf)
        img.width = 1200
        img.height = 326
        img.anchor = f"B{strip_anchor_row}"
        ws.add_image(img)

    # ====================================================================
    # SECTION 4 — TACTICAL ACTIONS TABLE
    # ====================================================================
    action_header_row = strip_anchor_row + 19
    ws.cell(row=action_header_row, column=2,
            value="TACTICAL ACTIONS — next 12 hours"
            ).font = HEADER_FONT
    ws.merge_cells(start_row=action_header_row, start_column=2,
                   end_row=action_header_row, end_column=17)

    # Header row: Segment | Time | Risk | Action
    hdr_row = action_header_row + 1
    headers = [("Segment", 2, 1), ("Time", 3, 3), ("Risk", 6, 1),
               ("Recommended action", 7, 11)]
    for label, start_col, span in headers:
        c = ws.cell(row=hdr_row, column=start_col, value=label)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = TITLE_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER
        if span > 1:
            ws.merge_cells(start_row=hdr_row, start_column=start_col,
                           end_row=hdr_row, end_column=start_col + span - 1)
        for col in range(start_col + 1, start_col + span):
            ws.cell(row=hdr_row, column=col).border = THIN_BORDER

    for i, seg in enumerate(segments):
        r = hdr_row + 1 + i
        ws.row_dimensions[r].height = 36

        # Segment label
        ws.cell(row=r, column=2, value=f"#{seg.idx + 1}").font = Font(
            name="Calibri", size=11, bold=True, color="333333")
        ws.cell(row=r, column=2).alignment = Alignment(
            horizontal="center", vertical="center")
        ws.cell(row=r, column=2).border = THIN_BORDER

        # Time
        time_text = seg.label
        ws.cell(row=r, column=3, value=time_text).font = ACTION_TEXT_FONT
        ws.cell(row=r, column=3).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        for c in range(3, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER

        # Risk label
        risk_cell = ws.cell(row=r, column=6, value=seg.risk_level.upper())
        risk_cell.font = ACTION_LABEL_FONT
        risk_cell.fill = RISK_LABEL_FILL.get(seg.risk_level, RISK_LABEL_FILL["green"])
        risk_cell.alignment = Alignment(horizontal="center", vertical="center")
        risk_cell.border = THIN_BORDER

        # Action text (wrapped)
        action_cell = ws.cell(row=r, column=7, value=seg.action)
        action_cell.font = ACTION_TEXT_FONT
        action_cell.fill = RISK_FILL.get(seg.risk_level, RISK_FILL["green"])
        action_cell.alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True, indent=1)
        action_cell.border = THIN_BORDER
        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=17)
        for c in range(8, 18):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).fill = RISK_FILL.get(seg.risk_level, RISK_FILL["green"])

    # ====================================================================
    # FOOTER — methodology + interpretation notes
    # ====================================================================
    notes_row = hdr_row + 2 + len(segments) + 1
    ws.cell(row=notes_row, column=2,
            value="How to read this tab").font = HEADER_FONT

    notes = [
        "• Window starts at this plan's depart_hour and covers the next 12 hours in 4 × 3-hour segments.",
        "• Each watch card shows forecast conditions (wind, seas, pressure) and cat-polar performance (boat speed) for that 3-hour segment.",
        "• Boat speed shows BOTH polar potential (clean theoretical Vs) AND calibrated expected (polar × sea_factor for the conditions). Calibrated is what you'll actually do.",
        "• Risk color matches the source Plan tab (green/yellow/red) — driven by gust thresholds, sea state, and night transitions.",
        "• Position label tells you which leg you're sailing during that segment.",
        "• Tactical actions are derived from sail-mode transitions, gust thresholds, sunrise/sunset, and convective probability.",
        "• Refresh this tab by rebuilding the workbook with a fresher forecast cycle and updated underway position.",
    ]
    for i, note in enumerate(notes):
        r = notes_row + 1 + i
        c = ws.cell(row=r, column=2, value=note)
        c.font = Font(name="Calibri", size=9, color="555555")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=17)
