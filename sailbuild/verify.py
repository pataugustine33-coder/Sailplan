"""
verify.py — systematic end-of-run cell-by-cell verification for sailbuild workbooks.

Runs after every build to catch the kinds of bugs that real navigators notice
but a passing test suite wouldn't flag: hard-coded vessel references that leak
into the wrong workbook, polar grid mismatches between Vessel Particulars and
compute.py, buoy coordinate mismatches against NDBC, wrong-vessel metrics in
the Risk Bowtie, and arrival timing that's slipped into the night window.

Each check returns a Finding(severity, location, message).
severity: 'error' (build is wrong, fix before delivery)
          'warn'  (likely correct but worth a manual look)
          'info'  (note for the skipper)

Usage from build.py:
    from sailbuild.verify import verify_workbook
    findings = verify_workbook(output_path, passage, forecast, buoys)
    for f in findings:
        print(f.format())
"""
from __future__ import annotations
from dataclasses import dataclass
from typing import Optional
from openpyxl import load_workbook


@dataclass
class Finding:
    severity: str  # 'error' | 'warn' | 'info'
    location: str  # 'tab:cell' or 'YAML:path'
    message: str

    def format(self) -> str:
        sev_icon = {"error": "❌", "warn": "⚠️ ", "info": "ℹ️ "}.get(self.severity, "  ")
        return f"  {sev_icon} [{self.severity.upper()}] {self.location}: {self.message}"


def verify_workbook(output_path: str, passage: dict, forecast: dict, buoys: dict) -> list[Finding]:
    """Run all verification checks against a built workbook.

    Returns a list of findings ordered by severity (errors first).
    Empty list means clean build.
    """
    wb = load_workbook(output_path)
    findings: list[Finding] = []

    findings.extend(_check_vessel_consistency(wb, passage))
    findings.extend(_check_polar_consistency(wb, passage))
    findings.extend(_check_buoy_coords_against_yaml(wb, buoys))
    findings.extend(_check_arrival_timing(wb, passage))
    findings.extend(_check_no_hr48_leakage(wb, passage))
    findings.extend(_check_tab_inventory(wb, passage))
    findings.extend(_check_forecast_freshness(wb, forecast))
    findings.extend(_check_gust_data_populated(wb, passage, forecast))
    findings.extend(_check_wp_zone_geography(passage, forecast))
    findings.extend(_check_plan_tab_images(wb, passage, forecast))
    findings.extend(_check_workbook_cells(wb, forecast))
    findings.extend(_check_methodology_compliance(wb, passage, forecast))

    # Sort: errors first, then warnings, then info
    severity_order = {"error": 0, "warn": 1, "info": 2}
    findings.sort(key=lambda f: severity_order.get(f.severity, 99))
    return findings


# ==============================================================
# Individual checks
# ==============================================================

def _check_vessel_consistency(wb, passage: dict) -> list[Finding]:
    """The vessel designation in passage YAML should appear in tab titles
    that reference the vessel, and the opposite designation should NOT.

    Catches: HR 48 references that leak into HR 54 workbooks (and vice versa),
    and brand/designer terms (Frers, Hallberg-Rassy) that leak into non-HR
    vessel workbooks (e.g. an Oyster 475 build that still says "Frers VPP" on
    its polar grid header — a real bug seen 5/19).
    """
    findings = []
    vessel = passage.get("vessel", {})
    vessel_label = vessel.get("designation", "")
    if not vessel_label:
        return findings

    # Determine the "other" common designations that should NOT appear in this
    # workbook unless the Vessel Comparison tab is intentionally included.
    other_designations = []
    if "48" in vessel_label:
        other_designations = ["HR 54", "HR54"]
    elif "54" in vessel_label:
        other_designations = ["HR 48", "HR48"]

    # Brand-leak terms: if the vessel YAML specifies a non-HR builder, then
    # "Hallberg-Rassy" and "Frers" should NOT appear anywhere in the workbook
    # outside the Vessel Comparison tab. Conversely, an HR build is allowed
    # to mention Hallberg-Rassy and Frers.
    builder = (vessel.get("builder") or "").lower()
    designer = (vessel.get("designer") or "").lower()
    is_hr_vessel = "hallberg" in builder or "rassy" in builder or vessel_label.upper().startswith("HR")
    if not is_hr_vessel:
        other_designations.extend(["Hallberg-Rassy", "Hallberg Rassy", "Frers", "Germán Frers", "German Frers"])

    # Skip the Vessel Comparison tab — it's allowed to mention both.
    skip_tabs = {"Vessel Comparison"}

    for tab in wb.sheetnames:
        if tab in skip_tabs:
            continue
        ws = wb[tab]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None or not isinstance(cell.value, str):
                    continue
                for other in other_designations:
                    if other in cell.value:
                        findings.append(Finding(
                            "error",
                            f"{tab}:{cell.coordinate}",
                            f"Cell mentions '{other}' but this passage is '{vessel_label}'. "
                            f"Cell content: {cell.value[:80]!r}",
                        ))
    return findings


def _check_polar_consistency(wb, passage: dict) -> list[Finding]:
    """Vessel Particulars polar grid header should match the passage's design_number.

    Catches: VS_GRID_D1170 alias bug where Vessel Particulars showed HR48
    values even when the passage was HR54.
    """
    findings = []
    design = passage.get("vessel", {}).get("design_number", "")
    if not design or "Vessel Particulars" not in wb.sheetnames:
        return findings

    ws = wb["Vessel Particulars"]
    # Polar subtitle row varies by renderer version — search rows 3-7 for
    # any cell that mentions "Polar" or "TWA × TWS".
    subtitle_found = False
    polar_subtitle_row = None
    for r in range(3, 8):
        v = str(ws.cell(r, 1).value or "")
        if "Polar" in v and ("TWA" in v or "TWS" in v):
            subtitle_found = True
            polar_subtitle_row = r
            if design not in v:
                findings.append(Finding(
                    "error",
                    f"Vessel Particulars:A{r}",
                    f"Polar grid subtitle does not include design_number '{design}'. "
                    f"Got: {v[:80]!r}",
                ))
            break
    if not subtitle_found:
        findings.append(Finding(
            "warn",
            "Vessel Particulars:A3-A7",
            "Could not locate polar grid subtitle in rows 3-7.",
        ))

    # Spot-check polar values — find the TWS header row first (the row that
    # starts with "TWA" in column A and has "TWS" in adjacent columns).
    tws_header_row = None
    for r in range(3, 12):
        v = str(ws.cell(r, 1).value or "")
        if v.startswith("TWA"):
            # Check that col 2 has "TWS"
            if "TWS" in str(ws.cell(r, 2).value or ""):
                tws_header_row = r
                break
    if tws_header_row is None:
        return findings

    # TWA grid is [45, 52, 60, 70, 80, 90, ...] — find the TWA=90 row
    # and column TWS=10 (4th TWS column).
    expected = {"D1170": 7.93, "D1206": 8.23}.get(design)
    if expected is None:
        return findings
    for offset in range(1, 12):
        twa_val = ws.cell(tws_header_row + offset, 1).value
        if twa_val == 90 or twa_val == 90.0:
            # TWS columns start at col 2 with TWS=4,6,8,10 → TWS=10 is col 5
            val = ws.cell(tws_header_row + offset, 5).value
            if val is None:
                break
            if abs(float(val) - expected) > 0.05:
                findings.append(Finding(
                    "error",
                    f"Vessel Particulars:E{tws_header_row + offset}",
                    f"Polar value at TWA 90/TWS 10 is {val}, expected ~{expected} for {design}. "
                    f"Wrong polar grid likely embedded.",
                ))
            break

    return findings


def _check_buoy_coords_against_yaml(wb, buoys: dict) -> list[Finding]:
    """Buoy lat/lon shown in Live Buoy Data should match what's in the YAML.

    Note: This is a self-consistency check (YAML→workbook). For NDBC-truth
    verification, see verify_buoys_against_ndbc() which is a separate
    Chrome-fetched check (called separately when desired).
    """
    findings = []
    if "Live Buoy Data" not in wb.sheetnames or "stations" not in buoys:
        return findings

    # Build a quick id→(lat,lon) map from YAML
    yaml_coords = {
        str(st["id"]): (st.get("lat"), st.get("lon"))
        for st in buoys["stations"]
    }

    # Live Buoy Data renders station id in col 1 area; lat/lon are not always shown
    # in this tab — but if they are, they should match. This check is best-effort.
    # The more reliable check is the Buoys by WP tab.
    if "Buoys by WP" in wb.sheetnames:
        ws = wb["Buoys by WP"]
        for r in range(4, ws.max_row + 1):
            buoy_id = ws.cell(r, 2).value
            lat_str = ws.cell(r, 4).value
            lon_str = ws.cell(r, 5).value
            if buoy_id is None or lat_str is None or lon_str is None:
                continue
            buoy_id = str(buoy_id)
            try:
                lat_val = float(str(lat_str).rstrip("°N").rstrip("°S").rstrip())
                lon_val = float(str(lon_str).rstrip("°W").rstrip("°E").rstrip())
            except (ValueError, AttributeError):
                continue
            yaml_lat, yaml_lon = yaml_coords.get(buoy_id, (None, None))
            if yaml_lat is None:
                # Buoy in workbook but not in buoys YAML — note for cross-check
                continue
            if abs(lat_val - yaml_lat) > 0.01:
                findings.append(Finding(
                    "warn",
                    f"Buoys by WP:D{r}",
                    f"Buoy {buoy_id} lat {lat_val} differs from YAML {yaml_lat} by "
                    f">0.01° (~0.6 NM). Verify against NDBC station page.",
                ))
            if abs(abs(lon_val) - abs(yaml_lon)) > 0.01:
                findings.append(Finding(
                    "warn",
                    f"Buoys by WP:E{r}",
                    f"Buoy {buoy_id} lon {lon_val} differs from YAML {yaml_lon} by "
                    f">0.01° (~0.6 NM). Verify against NDBC station page.",
                ))
    return findings


def _check_arrival_timing(wb, passage: dict) -> list[Finding]:
    """Arrival ETA color should match daylight window in passage.arrival_timing.

    Catches: yellow/red arrival not flagged in the briefing.

    Plan tabs have an ETA in column D for each WP row, but ALSO have a
    position-legend section further down that uses column D for descriptive
    text ("Starboard Bow — ..."). We must stop at the "PASSAGE TOTALS" row
    which marks the end of the WP data block.
    """
    findings = []
    arrival_timing = passage.get("arrival_timing", {})
    if not arrival_timing:
        return findings

    for tab_name in wb.sheetnames:
        if not tab_name.startswith("Plan ") or tab_name.endswith("Bowtie"):
            continue
        ws = wb[tab_name]
        # Find the totals marker (col B says "PASSAGE TOTALS") to bound the WP block.
        totals_row = None
        for r in range(4, ws.max_row + 1):
            b_val = ws.cell(r, 2).value
            if b_val and "PASSAGE TOTAL" in str(b_val).upper():
                totals_row = r
                break

        # Final WP arrival is the LAST row before totals where col D has a value
        # AND the value looks like a time (contains "AM" or "PM").
        upper_bound = totals_row if totals_row else ws.max_row + 1
        last_eta_row = None
        for r in range(4, upper_bound):
            d_val = ws.cell(r, 4).value
            if d_val and ("AM" in str(d_val) or "PM" in str(d_val)):
                last_eta_row = r

        if last_eta_row is None:
            continue
        eta_cell = ws.cell(last_eta_row, 4)
        try:
            fill_color = eta_cell.fill.start_color.value or ""
        except AttributeError:
            fill_color = ""
        eta_text = ws.cell(last_eta_row, 4).value
        if "FFC7CE" in str(fill_color).upper():
            findings.append(Finding(
                "error",
                f"{tab_name}:D{last_eta_row}",
                f"Arrival ETA {eta_text!r} is in NIGHT window (red). "
                f"Consider departure shift using reverse calculator on Format Reference tab.",
            ))
        elif "FFEB9C" in str(fill_color).upper():
            findings.append(Finding(
                "warn",
                f"{tab_name}:D{last_eta_row}",
                f"Arrival ETA {eta_text!r} is in TWILIGHT window (yellow). Marginal — "
                f"verify daylight margin and channel-marker lighting at destination.",
            ))
    return findings


def _check_no_hr48_leakage(wb, passage: dict) -> list[Finding]:
    """Scan every cell for known hard-coded HR48 phrases that should have been
    parameterized. This is the canary check that flags any regression.
    """
    findings = []
    vessel_label = passage.get("vessel", {}).get("designation", "")
    if "48" in vessel_label:
        return []  # this IS HR 48, so HR 48 mentions are correct

    # Known leak phrases that historically appeared hard-coded in renderers.
    # These should ALL be parameterized via vessel.designation now.
    canary_phrases = [
        "manageable for HR 48",
        "HR 48 Comfort & Safety Radar",
        "HR 48 Mk II against this plan",
        "HR 48 polar(",
    ]

    skip_tabs = {"Vessel Comparison"}
    for tab in wb.sheetnames:
        if tab in skip_tabs:
            continue
        ws = wb[tab]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None or not isinstance(cell.value, str):
                    continue
                for phrase in canary_phrases:
                    if phrase in cell.value:
                        findings.append(Finding(
                            "error",
                            f"{tab}:{cell.coordinate}",
                            f"Hard-coded HR 48 phrase leaked into a {vessel_label} workbook: "
                            f"{phrase!r}. Renderer is not parameterized.",
                        ))
    return findings


def _check_tab_inventory(wb, passage: dict) -> list[Finding]:
    """Tab count matches expected based on number of plans and Vessel Comparison flag."""
    findings = []
    n_plans = len(passage.get("plans", []))
    include_comparison = passage.get("include_vessel_comparison", False)

    # Base tabs: Briefing, Format Reference, Waypoints, Vessel Particulars,
    # Live Buoy Data, Refresh Cadence, Forecast Sources by WP, Buoys by WP,
    # Forecast Products, URL Quick Reference, Glossary, Verification Scorecard
    # = 12 fixed
    expected = 12
    expected += 2 * n_plans  # Plan + Bowtie per plan
    if include_comparison:
        expected += 1

    actual = len(wb.sheetnames)
    if actual != expected:
        findings.append(Finding(
            "warn",
            "workbook:tabs",
            f"Tab count is {actual}, expected {expected} "
            f"(12 base + {2*n_plans} plan/bowtie pairs"
            f"{' + 1 Vessel Comparison' if include_comparison else ''}).",
        ))
    return findings


def _check_gust_data_populated(wb, passage: dict, forecast: dict) -> list[Finding]:
    """Plan tabs: column I (Gust kt) must not be "—" for WPs whose forecast
    zone-period has `wind_gust_kt` defined in the YAML.

    Catches: regression where the column is rendered but the value isn't
    written, leaving the navigator looking at a forecast without gust data
    even though the YAML had it. Gusts are decision-relevant for reefing.
    """
    findings = []
    zones = forecast.get("zones", {})
    assignments_by_plan = forecast.get("waypoint_assignments", {})

    # Build tab_label → plan_id map so we check each Plan tab against ITS
    # plan's assignments, not all plans'.
    tab_to_plan_id = {}
    for plan in passage.get("plans", []):
        label = plan.get("tab_label", "")
        # Tab name may be truncated to 31 chars (openpyxl limit); match by prefix
        tab_to_plan_id[label[:31]] = plan["id"]

    for tab_name in wb.sheetnames:
        if not tab_name.startswith("Plan ") or tab_name.endswith("Bowtie"):
            continue
        plan_id = tab_to_plan_id.get(tab_name)
        if plan_id is None:
            continue  # not a plan tab we can map (defensive)
        plan_assignments = assignments_by_plan.get(plan_id, {})
        ws = wb[tab_name]

        # Walk WP rows (start at row 4, stop at PASSAGE TOTALS marker).
        totals_row = None
        for r in range(4, ws.max_row + 1):
            b_val = ws.cell(r, 2).value
            if b_val and "PASSAGE TOTAL" in str(b_val).upper():
                totals_row = r
                break
        upper = totals_row if totals_row else ws.max_row + 1

        for r in range(4, upper):
            wp_id = ws.cell(r, 1).value
            gust_cell = ws.cell(r, 9).value
            if not wp_id or wp_id == "WP":
                continue
            if str(wp_id) not in plan_assignments:
                continue
            assignment = plan_assignments[str(wp_id)]
            zone = assignment.get("zone")
            period = assignment.get("period")
            if not zone or not period:
                continue
            period_data = zones.get(zone, {}).get("periods", {}).get(period, {})
            yaml_gust = period_data.get("wind_gust_kt")
            if yaml_gust and (gust_cell == "—" or gust_cell is None):
                findings.append(Finding(
                    "error",
                    f"{tab_name}:I{r}",
                    f"Gust kt cell is empty/dash but YAML "
                    f"({zone} {period} for {plan_id}) specifies wind_gust_kt={yaml_gust}. "
                    f"Renderer is dropping gust data on the floor.",
                ))

        # Second check: if NONE of the WPs in this plan have gust data populated,
        # warn — the forecast YAML likely doesn't have wind_gust_kt for any of
        # the zone-periods used. NWS often omits gusts when conditions are benign,
        # but for operational use we usually want at least an estimated gust ceiling.
        gust_values_found = 0
        for r in range(4, upper):
            wp_id = ws.cell(r, 1).value
            gust_cell = ws.cell(r, 9).value
            if not wp_id or wp_id == "WP":
                continue
            if gust_cell and gust_cell != "—" and gust_cell is not None:
                gust_values_found += 1
        if gust_values_found == 0:
            findings.append(Finding(
                "warn",
                f"{tab_name}:I (column)",
                f"No gust data populated for any WP in this plan. Forecast YAML "
                f"may be missing wind_gust_kt for the zone-periods used. "
                f"Consider adding gust estimates from buoy observations.",
            ))
    return findings


def _check_forecast_freshness(wb, forecast: dict) -> list[Finding]:
    """Format Reference freshness panel should not show any sources as STALE."""
    findings = []
    if "Format Reference" not in wb.sheetnames:
        return findings
    ws = wb["Format Reference"]
    # Status column is column 5 (header row 4). Data rows start at 5.
    for r in range(5, 15):
        status = ws.cell(r, 5).value
        if status is None:
            break
        status_str = str(status).upper()
        if "STALE" in status_str or "EXPIRED" in status_str:
            source = ws.cell(r, 1).value
            findings.append(Finding(
                "warn",
                f"Format Reference:E{r}",
                f"Source {source!r} is {status_str}. Pull fresh cycle before departure.",
            ))
        elif "OFFLINE" in status_str:
            source = ws.cell(r, 1).value
            findings.append(Finding(
                "info",
                f"Format Reference:E{r}",
                f"Source {source!r} is OFFLINE. Cross-check with neighboring buoys.",
            ))
    return findings


# ==============================================================
# Geographic WP-to-zone validator
# ==============================================================
#
# Catches a class of YAML errors where a waypoint's `zone:` assignment
# doesn't match the WP's actual geographic position. This is silent
# corruption: pipeline runs clean, but the wind/sea data driving the
# polar model is wrong because it's pulled from the wrong zone's forecast.
#
# Strategy: each known marine zone is paired with one or more "coastline
# reference points" plus a distance band (inshore 0-20 NM, offshore
# 20-60 NM). For each WP, we Haversine to the nearest reference for its
# assigned zone and check that:
#   1. The WP's distance from nearest land falls in the zone's band.
#   2. The latitude band of the zone covers the WP.
#
# Reference data is intentionally over-broad — multiple plausible refs
# per zone — to avoid flagging legitimate variation in route geometry.

import math


def _haversine_nm(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Great-circle distance in nautical miles."""
    R_NM = 3440.065
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat / 2) ** 2
         + math.cos(math.radians(lat1))
         * math.cos(math.radians(lat2))
         * math.sin(dlon / 2) ** 2)
    return 2 * R_NM * math.asin(math.sqrt(a))


# Per-zone metadata for the SC/NC/FL east coast corridors this project
# operates in. `band_nm` is (min, max) distance offshore. `lat_band` is
# (south, north) latitude. `coast_refs` is points along the coastline
# segment the zone covers — we compute min distance to any of them.
#
# To extend to new zones, just add an entry. Unknown zones produce a
# WARN (not ERROR) so unrecognized zones don't block a build.
ZONE_REGISTRY = {
    # KCHS — Charleston SC
    "AMZ340": {
        "office": "CHS", "band_nm": (0, 20), "lat_band": (32.5, 33.0),
        "coast_refs": [(32.78, -79.92)],
        "description": "Charleston Harbor",
    },
    "AMZ360": {
        "office": "CHS", "band_nm": (0, 20), "lat_band": (32.4, 33.2),
        "coast_refs": [(32.49, -80.31), (32.78, -79.92), (33.13, -79.27)],
        "description": "South Santee River to Edisto Beach SC, out 20 NM",
    },
    "AMZ362": {
        "office": "CHS", "band_nm": (0, 20), "lat_band": (32.0, 32.6),
        "coast_refs": [(32.49, -80.31), (32.04, -80.86)],
        "description": "Edisto Beach SC to Savannah GA, out 20 NM",
    },
    "AMZ364": {
        "office": "CHS", "band_nm": (0, 20), "lat_band": (31.3, 32.1),
        "coast_refs": [(32.04, -80.86), (31.40, -81.30)],
        "description": "Savannah GA to Altamaha Sound GA, out 20 NM",
    },
    "AMZ380": {
        "office": "CHS", "band_nm": (20, 60), "lat_band": (32.4, 33.2),
        "coast_refs": [(32.49, -80.31), (32.78, -79.92), (33.13, -79.27)],
        "description": "South Santee River to Edisto Beach SC, 20-60 NM",
    },
    "AMZ382": {
        "office": "CHS", "band_nm": (20, 60), "lat_band": (32.0, 32.6),
        "coast_refs": [(32.49, -80.31), (32.04, -80.86)],
        "description": "Edisto Beach SC to Savannah GA, 20-60 NM",
    },
    "AMZ384": {
        "office": "CHS", "band_nm": (20, 60), "lat_band": (31.3, 32.1),
        "coast_refs": [(32.04, -80.86), (31.40, -81.30)],
        "description": "Savannah GA to Altamaha Sound GA, 20-60 NM",
    },
    # KILM — Wilmington NC
    "AMZ250": {
        "office": "ILM", "band_nm": (0, 20), "lat_band": (33.8, 34.5),
        "coast_refs": [(33.84, -78.01), (34.21, -77.79), (34.42, -77.55)],
        "description": "Surf City to Cape Fear NC, out 20 NM",
    },
    "AMZ252": {
        "office": "ILM", "band_nm": (0, 20), "lat_band": (33.6, 34.0),
        "coast_refs": [(33.84, -78.01), (33.85, -78.65)],
        "description": "Cape Fear NC to Little River Inlet SC, out 20 NM",
    },
    "AMZ254": {
        "office": "ILM", "band_nm": (0, 20), "lat_band": (33.5, 33.9),
        "coast_refs": [(33.85, -78.65), (33.66, -79.03)],
        "description": "Little River Inlet to Murrells Inlet SC, out 20 NM",
    },
    "AMZ256": {
        "office": "ILM", "band_nm": (0, 20), "lat_band": (33.1, 33.7),
        "coast_refs": [(33.51, -79.03), (33.13, -79.27)],
        "description": "Murrells Inlet to South Santee River SC, out 20 NM",
    },
    "AMZ280": {
        "office": "ILM", "band_nm": (20, 60), "lat_band": (33.1, 34.5),
        "coast_refs": [(34.42, -77.55), (34.21, -77.79), (33.84, -78.01),
                       (33.85, -78.65), (33.66, -79.03)],
        "description": "Surf City NC to Little River Inlet SC, 20-60 NM",
    },
    "AMZ284": {
        "office": "ILM", "band_nm": (20, 60), "lat_band": (33.1, 33.9),
        "coast_refs": [(33.85, -78.65), (33.66, -79.03), (33.13, -79.27)],
        "description": "Little River Inlet to South Santee River SC, 20-60 NM",
    },
    # KMHX — Newport/Morehead NC
    "AMZ150": {
        "office": "MHX", "band_nm": (0, 20), "lat_band": (35.7, 36.5),
        "coast_refs": [(35.79, -75.55), (36.13, -75.71)],
        "description": "Currituck Beach Light to Oregon Inlet NC, out 20 NM",
    },
    "AMZ152": {
        "office": "MHX", "band_nm": (0, 20), "lat_band": (35.2, 35.8),
        "coast_refs": [(35.27, -75.53), (35.79, -75.55)],
        "description": "Oregon Inlet to Cape Hatteras NC, out 20 NM",
    },
    "AMZ154": {
        "office": "MHX", "band_nm": (0, 20), "lat_band": (34.9, 35.3),
        "coast_refs": [(35.27, -75.53), (34.97, -76.04)],
        "description": "Cape Hatteras to Ocracoke Inlet NC, out 20 NM",
    },
    "AMZ156": {
        "office": "MHX", "band_nm": (0, 20), "lat_band": (34.5, 35.0),
        "coast_refs": [(34.97, -76.04), (34.62, -76.52)],
        "description": "Ocracoke Inlet to Cape Lookout NC, out 20 NM",
    },
    "AMZ158": {
        "office": "MHX", "band_nm": (0, 20), "lat_band": (34.3, 34.8),
        "coast_refs": [(34.62, -76.52), (34.42, -77.55)],
        "description": "Cape Lookout to Surf City NC, out 20 NM",
    },
    "AMZ188": {
        "office": "MHX", "band_nm": (20, 60), "lat_band": (33.9, 34.8),
        "coast_refs": [(34.62, -76.52), (34.42, -77.55)],
        "description": "Cape Lookout to Surf City NC, 20-60 NM",
    },
    # KMFL — Miami FL (SE Florida coast)
    "AMZ650": {
        "office": "MFL", "band_nm": (0, 20), "lat_band": (26.20, 26.95),
        "coast_refs": [(26.77, -80.04), (26.55, -80.06), (26.32, -80.08)],
        "description": "Jupiter Inlet to Deerfield Beach FL, out 20 NM",
    },
    "AMZ651": {
        "office": "MFL", "band_nm": (0, 20), "lat_band": (25.30, 26.35),
        "coast_refs": [(26.32, -80.08), (26.09, -80.10), (25.77, -80.13),
                       (25.34, -80.27)],
        "description": "Deerfield Beach to Ocean Reef FL, out 20 NM",
    },
    "AMZ670": {
        "office": "MFL", "band_nm": (20, 60), "lat_band": (26.20, 26.95),
        "coast_refs": [(26.77, -80.04), (26.55, -80.06), (26.32, -80.08)],
        "description": "Jupiter Inlet to Deerfield Beach FL, 20-60 NM",
    },
    "AMZ671": {
        "office": "MFL", "band_nm": (20, 60), "lat_band": (25.30, 26.35),
        "coast_refs": [(26.32, -80.08), (26.09, -80.10), (25.77, -80.13),
                       (25.34, -80.27)],
        "description": "Deerfield Beach to Ocean Reef FL, 20-60 NM",
    },
    # Add JAX/MLB/KEY zones as new routes are built. Unknown zones
    # produce a WARN, not an ERROR.
}


def _check_wp_zone_geography(passage: dict, forecast: dict) -> list[Finding]:
    """Verify each WP's assigned zone matches the WP's geographic position.

    Checks two things per WP:
      1. WP latitude falls within the zone's lat band (gross sanity check).
      2. WP distance from nearest coastline reference for the zone falls
         within the zone's distance band (inshore 0-20 NM vs offshore
         20-60 NM).

    Catches the silent-corruption class where a coastal-zone bulletin is
    used to drive the polar model at a waypoint that's actually in the
    offshore band — materially different forecast (and a real bug seen
    Tue 5/19).
    """
    findings: list[Finding] = []

    wp_by_id = {wp["id"]: wp for wp in passage.get("waypoints", [])}
    assignments = forecast.get("waypoint_assignments", {})

    for plan_id, plan_wps in assignments.items():
        for wp_id, assignment in plan_wps.items():
            zone = assignment.get("zone")
            if not zone:
                findings.append(Finding(
                    "error",
                    f"forecast YAML:waypoint_assignments.{plan_id}.{wp_id}",
                    "missing zone: assignment",
                ))
                continue

            wp = wp_by_id.get(wp_id)
            if not wp:
                findings.append(Finding(
                    "error",
                    f"forecast YAML:waypoint_assignments.{plan_id}.{wp_id}",
                    f"WP {wp_id!r} assigned a zone but not defined in passage.waypoints",
                ))
                continue

            wp_lat = wp.get("lat")
            wp_lon = wp.get("lon")
            if wp_lat is None or wp_lon is None:
                continue

            zone_info = ZONE_REGISTRY.get(zone)
            if not zone_info:
                findings.append(Finding(
                    "warn",
                    f"forecast YAML:waypoint_assignments.{plan_id}.{wp_id}",
                    f"Zone {zone!r} not in geographic registry — "
                    f"cannot verify against {wp_id} position "
                    f"({wp_lat:.4f}, {wp_lon:.4f}). Add to ZONE_REGISTRY "
                    f"in sailbuild/verify.py to enable check.",
                ))
                continue

            # Lat band check
            lat_lo, lat_hi = zone_info["lat_band"]
            if not (lat_lo - 0.1 <= wp_lat <= lat_hi + 0.1):
                findings.append(Finding(
                    "error",
                    f"forecast YAML:waypoint_assignments.{plan_id}.{wp_id}",
                    f"{wp_id} lat {wp_lat:.4f} is outside zone {zone}'s "
                    f"lat band ({lat_lo:.2f}-{lat_hi:.2f}). Zone covers "
                    f"{zone_info['description']!r}. Wrong zone assigned.",
                ))
                continue

            # Distance band check
            min_dist = min(
                _haversine_nm(wp_lat, wp_lon, rlat, rlon)
                for rlat, rlon in zone_info["coast_refs"]
            )
            band_lo, band_hi = zone_info["band_nm"]
            # Allow a small tolerance at the boundary (±2 NM)
            if min_dist < band_lo - 2:
                findings.append(Finding(
                    "error",
                    f"forecast YAML:waypoint_assignments.{plan_id}.{wp_id}",
                    f"{wp_id} is {min_dist:.1f} NM from nearest coast point "
                    f"of zone {zone}, but the zone covers {band_lo}-{band_hi} NM "
                    f"offshore. WP is INSHORE of this zone — reassign to a "
                    f"0-20 NM zone in the same coastline segment.",
                ))
            elif min_dist > band_hi + 5:
                findings.append(Finding(
                    "error",
                    f"forecast YAML:waypoint_assignments.{plan_id}.{wp_id}",
                    f"{wp_id} is {min_dist:.1f} NM from nearest coast point "
                    f"of zone {zone}, but the zone covers only {band_lo}-{band_hi} NM "
                    f"offshore. WP is OUTSIDE this zone — reassign to a "
                    f"20-60 NM offshore zone or an OWF zone.",
                ))

    return findings


# ==============================================================
# Workbook cell scan — catches Excel errors, un-evaluated formulas,
# forbidden risk-label words, stale dates
# ==============================================================

import re as _re


# ==============================================================
# Image-presence check — catches missing wind/sea roses, polars,
# and timeline strips on plan tabs
# ==============================================================
#
# This catches a class of failure where build.py runs clean and the
# workbook looks valid (cells contain "—" placeholder) but a critical
# rendering library was missing in the build environment and the chart
# silently fell back to a dash. Real example: cairosvg not installed
# → every Wind/Sea Rose became "—" while the cell-scan check happily
# saw "—" as a non-empty string and moved on.

def _check_plan_tab_images(wb, passage: dict, forecast: dict) -> list[Finding]:
    """Verify each plan tab has the expected count of embedded images.

    Each plan tab should carry:
      - One wind/sea rose per non-arrival WP with a forecast assignment
        (anchored to column U)
      - One mini-polar per non-arrival WP with a forecast assignment
        (anchored to column V)
      - One timeline strip below the leg table

    Anchor columns are 0-indexed in openpyxl's image API. U=20, V=21.
    """
    findings: list[Finding] = []

    waypoints = passage.get("waypoints", [])
    if not waypoints:
        return findings

    # Number of WPs that should have rose/polar imagery rendered.
    # Two conditions must hold:
    #   1. WP has course_out (i.e., outbound leg exists) — terminal WPs don't.
    #   2. WP is referenced in this plan's waypoint_assignments —
    #      unassigned WPs (most commonly WP0 the departure point) render
    #      as no-leg rows and skip the image embed.
    # Counting both keeps this check honest across passages that do vs
    # don't include WP0 in assignments.
    plans = passage.get("plans", [])
    wp_assignments = forecast.get("waypoint_assignments", {}) or {}

    for plan_def in plans:
        tab_name = plan_def.get("tab_label")
        if not tab_name or tab_name not in wb.sheetnames:
            continue
        ws = wb[tab_name]

        # Per-plan expected count: WPs that have BOTH course_out (outbound
        # leg) AND a forecast assignment in this plan.
        plan_assignments = wp_assignments.get(plan_def["id"], {}) or {}
        expected_per_col = sum(
            1 for wp in waypoints
            if wp.get("course_out") is not None and wp["id"] in plan_assignments
        )

        rose_count = 0   # col U
        polar_count = 0  # col V
        other_count = 0
        for img in ws._images:
            anchor = img.anchor
            if not hasattr(anchor, "_from"):
                other_count += 1
                continue
            col = anchor._from.col
            if col == 20:    # U (Wind/Sea Rose)
                rose_count += 1
            elif col == 21:  # V (Polar @ TWS)
                polar_count += 1
            else:
                other_count += 1

        if rose_count < expected_per_col:
            findings.append(Finding(
                "error",
                f"{tab_name}:U (Wind/Sea Rose column)",
                f"Only {rose_count}/{expected_per_col} wind/sea rose images "
                f"embedded. Likely cause: cairosvg not installed in the "
                f"build environment (rose renderer falls back to '—' "
                f"placeholder when HAVE_CAIROSVG is False). "
                f"Run: pip install cairosvg",
            ))
        if polar_count < expected_per_col:
            findings.append(Finding(
                "error",
                f"{tab_name}:V (Polar @ TWS column)",
                f"Only {polar_count}/{expected_per_col} mini-polar images "
                f"embedded. Check matplotlib/PIL availability in build "
                f"environment and inspect sailbuild/charts.py renderer.",
            ))
        if other_count == 0:
            findings.append(Finding(
                "warn",
                f"{tab_name}",
                "No timeline-strip image found below leg table. Check "
                "render_timeline_strip in sailbuild/tabs/plan.py.",
            ))

    return findings


_EXCEL_ERROR_LITERALS = {
    "#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NUM!",
    "#NULL!", "#GETTING_DATA",
}

# Risk cells (Weather Risk column on plan tabs) must be color-only.
# A leading color word violates the project standard.
_FORBIDDEN_RISK_LABEL_LEAD = _re.compile(r"^\s*(GREEN|YELLOW|RED)\b", _re.IGNORECASE)

# Placeholder markers indicate incomplete content.
_PLACEHOLDER_MARKERS = _re.compile(r"\b(TBC|TODO|FIXME|XXX|TBD|PLACEHOLDER)\b")


def _check_workbook_cells(wb, forecast: dict) -> list[Finding]:
    """Walk every populated cell on every tab and flag:

      - Excel formula errors (#REF!, #DIV/0!, #VALUE!, #N/A, etc).
      - Formulas that didn't evaluate (data_only view returned the
        formula string itself or None).
      - Risk-column cells that start with a color word (color-only
        rule from project standing prefs).
      - Placeholder markers (TBC / TODO / FIXME).
      - Date references that don't match the forecast cycle date,
        which usually means a stale string was carried over from a
        prior build's YAML.
    """
    findings: list[Finding] = []

    # Determine the build's "good" date strings so we can flag stale others.
    # Pull the cycle dates from the forecast YAML's cwfchs.issued ISO date.
    good_dates: set[str] = set()
    cycle = forecast.get("cycle", {}) if isinstance(forecast, dict) else {}
    for key in ("cwfchs", "cwfilm", "cwfmhx", "afdchs", "afdilm", "afdmhx"):
        product = cycle.get(key, {})
        issued = product.get("issued", "")
        # Pull YYYY-MM-DD prefix
        m = _re.match(r"(\d{4})-(\d{2})-(\d{2})", issued)
        if m:
            yyyy, mm, dd = m.groups()
            good_dates.add(f"{int(mm)}/{int(dd)}")
            good_dates.add(f"{int(mm)}/{int(dd)}/{yyyy[-2:]}")

    # If we couldn't extract the cycle date, skip the stale-date check
    # rather than emit false positives.
    do_stale_date_check = len(good_dates) > 0

    # Load a formula-view of the same workbook so we can detect
    # un-evaluated formulas.
    from openpyxl import load_workbook as _lwb
    wb_path = wb.path if hasattr(wb, "path") else None
    # openpyxl Workbook objects don't carry their source path back. The
    # caller passes data_only=True; we re-open with data_only=False here
    # by getting the path via the wb's loaded_from attribute pattern.
    # Simpler: just check formulas on `wb`. If the value is a string
    # starting with "=", it's an un-evaluated formula in this view.

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        is_plan = "Plan" in ws_name and "Bowtie" not in ws_name

        # Locate Weather Risk column on plan tabs (always col Q = 17 in
        # the current schema, but detect by header for robustness).
        risk_col = None
        if is_plan:
            for r in range(1, min(6, ws.max_row + 1)):
                for c in range(1, ws.max_column + 1):
                    if ws.cell(r, c).value == "Weather Risk":
                        risk_col = c
                        break
                if risk_col:
                    break

        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                if not isinstance(v, str):
                    continue

                # Excel error literals
                for err in _EXCEL_ERROR_LITERALS:
                    if err in v:
                        findings.append(Finding(
                            "error",
                            f"{ws_name}:{cell.coordinate}",
                            f"Cell contains Excel error literal {err!r}",
                        ))

                # Un-evaluated formula
                if v.startswith("="):
                    findings.append(Finding(
                        "error",
                        f"{ws_name}:{cell.coordinate}",
                        f"Formula not evaluated by LibreOffice "
                        f"(saw {v[:50]!r}). Rebuild with recalc.",
                    ))

                # Forbidden risk-cell color label
                if is_plan and cell.column == risk_col:
                    m = _FORBIDDEN_RISK_LABEL_LEAD.match(v)
                    if m:
                        findings.append(Finding(
                            "error",
                            f"{ws_name}:{cell.coordinate}",
                            f"Risk cell starts with color word {m.group(1)!r}. "
                            f"Standard says color-only fill, reason text "
                            f"without level label.",
                        ))

                # Placeholder markers
                m = _PLACEHOLDER_MARKERS.search(v)
                if m and len(v) < 200:
                    findings.append(Finding(
                        "warn",
                        f"{ws_name}:{cell.coordinate}",
                        f"Placeholder marker {m.group(0)!r} found in cell — "
                        f"content may be incomplete: {v[:80]!r}",
                    ))

                # Stale-date check: look for M/D patterns that aren't
                # in our good_dates set. Only flag if the cell looks
                # like a date label (contains a 3-letter day prefix
                # like Mon/Tue/Wed/Thu/Fri/Sat/Sun) AND the cell isn't
                # narrative text citing a historical example.
                if do_stale_date_check:
                    # Skip narrative lessons-learned cells that legitimately
                    # cite past cycles as examples or precedent.
                    is_historical_ref = bool(_re.search(
                        r"\b(example|cite|historical|previous|earlier|"
                        r"prior|past|precedent|lesson|reference|"
                        r"e\.?g\.?|i\.?e\.?|for instance|such as)\b",
                        v, _re.IGNORECASE,
                    ))
                    # Skip cells that are clearly narrative paragraphs
                    # rather than data — heuristic: more than 200 chars
                    # and contains complete sentences (period+space).
                    is_narrative = len(v) > 200 and ". " in v

                    if not (is_historical_ref or is_narrative):
                        for m in _re.finditer(r"\b(\d{1,2})/(\d{1,2})\b", v):
                            date_str = f"{int(m.group(1))}/{int(m.group(2))}"
                            if date_str in good_dates:
                                continue
                            ctx_start = max(0, m.start() - 25)
                            ctx = v[ctx_start:m.start()]
                            if _re.search(r"\b(Mon|Tue|Wed|Thu|Fri|Sat|Sun)\b", ctx):
                                findings.append(Finding(
                                    "warn",
                                    f"{ws_name}:{cell.coordinate}",
                                    f"Possible stale date {date_str!r} (build "
                                    f"cycle dates: {sorted(good_dates)}): "
                                    f"{v[:80]!r}",
                                ))
                                break  # one finding per cell

    return findings


# ==============================================================
# Methodology compliance check
# ==============================================================
#
# This is the catch-all "does the workbook deliver what the project
# methodology says it should?" check. The other verify functions cover
# specific failure modes (vessel consistency, polar grid match, buoy
# coord drift, image presence, etc). THIS function covers the structural
# contract: does every mandatory tab exist? Do plan tabs have every
# mandatory column? Do format conventions hold across the workbook?
#
# The rule set is declarative — represented as a METHODOLOGY constant —
# so when the standing instructions evolve (new column, new tab, new
# convention), you add a rule entry rather than write a new function.
#
# A rule that fails produces an `error` finding so the build's exit
# code reflects the violation and a downstream CI gate can block the
# delivery.
#
# Background: this check was added after a Tue 5/19 build incident
# where the assistant produced a hand-rolled workbook that bypassed
# build.py entirely. The hand-rolled output had 10 tabs and 17 plan
# columns; the project methodology says 16 tabs and 28 plan columns
# (17 mandatory A-Q + Sea Source + image columns + cum-time columns).
# A check that runs against the produced file and refuses to pass
# when the structural contract is broken makes this failure mode
# detectable at build time regardless of how the file was produced.


# Mandatory plan-tab columns per the project methodology.
# Order: required position. Type indicates how to verify content.
# 'mandatory' columns must be present with the exact header label.
# 'mandatory_with_alt' accepts the listed alternatives (older builds).
_PLAN_TAB_REQUIRED_COLUMNS = [
    # Spec'd in system prompt §"Plan tabs use 17 standard columns (A through Q)"
    ("WP",              "A",  ["WP"]),
    ("Description",     "B",  ["Description"]),
    ("Cum NM",          "C",  ["Cum NM"]),
    ("ETA (EDT)",       "D",  ["ETA (EDT)", "ETA"]),
    ("Pressure (inHg)", "E",  ["Pressure (inHg)", "Pressure"]),
    ("Pressure Trend",  "F",  ["Pressure Trend"]),
    ("Wind Dir",        "G",  ["Wind Dir"]),
    ("Wind kt",         "H",  ["Wind kt"]),
    ("Gust kt",         "I",  ["Gust kt"]),
    ("Sea Ht (ft)",     "J",  ["Sea Ht (ft)", "Sea Ht"]),
    ("Period (sec)",    "K",  ["Period (sec)", "Period"]),
    ("Course (°T)",     "L",  ["Course (°T)", "Course"]),
    ("TWA (°)",         "M",  ["TWA (°)", "TWA"]),
    # System prompt names "Sea Angle (°)"; current schema splits this
    # into "Sea From (°)" + "Sea Angle / Pos". Accept either form.
    ("Sea Angle (°)",   "N",  ["Sea Angle (°)", "Sea From (°)", "Sea Angle / Pos"]),
    ("Sail Mode",       "O",  ["Sail Mode"]),
    ("Notes",           "P",  ["Notes"]),
    ("Weather Risk",    "Q",  ["Weather Risk"]),
    # Standing instruction: Sea Source as column R. Position may have
    # shifted right in the current schema due to chart columns U/V; we
    # check presence-by-name, not strict position.
    ("Sea Source",      "R*", ["Sea Source"]),
]


# Mandatory workbook tabs per the project methodology.
# Each entry: (tab_name_or_predicate, severity_if_missing, why)
def _required_tabs(passage: dict) -> list[tuple]:
    """Build the list of required tab names for this passage.

    Plan and Bowtie tabs are dynamic (one per plan). Vessel Comparison
    is opt-in. Everything else is fixed.
    """
    base = [
        ("Pre-Departure Briefing", "error", "Top-level briefing with both plans side-by-side"),
        ("Format Reference",       "error", "Column spec, calibration, freshness panel, reverse calc"),
        ("Waypoints",              "error", "WP1+ with Lat/Lon/Cum NM/Leg NM/Course/Notes + map block"),
        ("Vessel Particulars",     "error", "Polar grid + stability + calibration for the modeled vessel"),
        ("Live Buoy Data",         "error", "NDBC obs aligned to forecast cycle for verification"),
        ("Forecast Sources by WP", "error", "Attribution trail: which zone covers which WP"),
        ("Buoys by WP",            "error", "Per-WP nearest buoy with distance + relevance"),
        ("Refresh Cadence",        "error", "Underway verification trigger schedule"),
        ("Forecast Products",      "error", "Catalog of CWF/AFD/OSO products with issuance times"),
        ("URL Quick Reference",    "error", "Live URLs used by this build for chrome-paste refresh"),
        ("Glossary",               "error", "Definitions: TWA, Sea Angle, Pressure Trend vocab, etc"),
        ("Verification Scorecard", "error", "Pre/under/post scorecard for verification methodology"),
    ]
    # Plan and Bowtie tabs, one per plan
    for plan in passage.get("plans", []):
        label = plan.get("tab_label")
        if label:
            base.append((label, "error", f"Plan tab for {plan.get('id', '?')}"))
            base.append((f"{label} Bowtie", "error",
                         f"Risk bowtie for {plan.get('id', '?')}"))
    if passage.get("include_vessel_comparison", False):
        base.append(("Vessel Comparison", "error", "HR48 vs HR54 side-by-side comparison"))
    return base


def _check_methodology_compliance(wb, passage: dict, forecast: dict) -> list[Finding]:
    """Verify the workbook delivers what the project methodology says it should.

    Runs five sub-checks:
      A. Mandatory tabs present
      B. Plan tabs have all mandatory columns (with header label match)
      C. Plan tabs have one row per WP (matching passage.waypoints)
      D. Pressure Trend values use the controlled vocabulary
      E. ETAs use the 12-hour day-prefix format ("Tue 3:00 PM")
    """
    findings: list[Finding] = []

    # --- A. Mandatory tabs ---
    required = _required_tabs(passage)
    present = set(wb.sheetnames)
    for tab_name, severity, why in required:
        if tab_name not in present:
            findings.append(Finding(
                severity,
                f"workbook:tabs",
                f"Missing mandatory tab {tab_name!r}. Purpose: {why}. "
                f"Project methodology requires this tab on every passage build.",
            ))

    # --- B. Plan-tab columns ---
    plan_tabs = [p.get("tab_label") for p in passage.get("plans", [])
                 if p.get("tab_label") in present]
    for tab in plan_tabs:
        ws = wb[tab]
        # Find header row (first row containing 'WP' in column 1)
        header_row = None
        for r in range(1, min(8, ws.max_row + 1)):
            if ws.cell(r, 1).value == "WP":
                header_row = r
                break
        if header_row is None:
            findings.append(Finding(
                "error",
                f"{tab}:A1",
                "No header row with 'WP' in column A found in first 7 rows. "
                "Plan tab structure is not recognizable.",
            ))
            continue

        headers = [ws.cell(header_row, c).value
                   for c in range(1, ws.max_column + 1)]

        for col_label, col_pos, alternatives in _PLAN_TAB_REQUIRED_COLUMNS:
            found = any(h in alternatives for h in headers if h is not None)
            if not found:
                findings.append(Finding(
                    "error",
                    f"{tab}:{col_pos}{header_row}",
                    f"Mandatory column {col_label!r} missing. "
                    f"Accepted header labels: {alternatives}. "
                    f"Project methodology specifies plan tabs must have this "
                    f"column at position {col_pos}.",
                ))

        # --- C. One row per WP, matching the passage waypoints ---
        expected_wps = [wp["id"] for wp in passage.get("waypoints", [])]
        wp_rows = {}
        for r in range(header_row + 1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if isinstance(v, str) and v.startswith("WP"):
                wp_rows[v] = r
        for wp_id in expected_wps:
            if wp_id not in wp_rows:
                findings.append(Finding(
                    "error",
                    f"{tab}:A (WP column)",
                    f"Plan tab missing row for {wp_id}. Passage defines "
                    f"{len(expected_wps)} waypoints; plan tab has "
                    f"{len(wp_rows)} WP rows.",
                ))

    # --- D. Pressure Trend vocabulary ---
    # Per system prompt: Rising fast / Rising / Rising slow / Steady /
    # Falling slow / Falling / Falling fast / Bottoming
    pressure_trend_vocab = {
        "Rising fast", "Rising", "Rising slow", "Steady",
        "Falling slow", "Falling", "Falling fast", "Bottoming",
    }
    for tab in plan_tabs:
        ws = wb[tab]
        header_row = None
        for r in range(1, 8):
            if ws.cell(r, 1).value == "WP":
                header_row = r
                break
        if header_row is None:
            continue
        headers = [ws.cell(header_row, c).value
                   for c in range(1, ws.max_column + 1)]
        if "Pressure Trend" not in headers:
            continue
        pt_col = headers.index("Pressure Trend") + 1
        for r in range(header_row + 1, ws.max_row + 1):
            wp = ws.cell(r, 1).value
            if not (isinstance(wp, str) and wp.startswith("WP")):
                continue
            pt = ws.cell(r, pt_col).value
            if pt is None or pt == "" or pt == "—":
                continue
            if isinstance(pt, str) and pt.strip() not in pressure_trend_vocab:
                findings.append(Finding(
                    "error",
                    f"{tab}:{_col_letter(pt_col)}{r}",
                    f"Pressure Trend {pt!r} is not in the project vocabulary. "
                    f"Allowed: {sorted(pressure_trend_vocab)}",
                ))

    # --- E. ETA format check ---
    # Per system prompt: 12-hour clock with day prefix, e.g. "Mon 1:50 PM"
    import re as _r
    eta_pattern = _r.compile(
        r"^(Mon|Tue|Wed|Thu|Fri|Sat|Sun) \d{1,2}:\d{2} (AM|PM)$"
    )
    for tab in plan_tabs:
        ws = wb[tab]
        header_row = None
        for r in range(1, 8):
            if ws.cell(r, 1).value == "WP":
                header_row = r
                break
        if header_row is None:
            continue
        headers = [ws.cell(header_row, c).value
                   for c in range(1, ws.max_column + 1)]
        eta_col = None
        for h in ("ETA (EDT)", "ETA"):
            if h in headers:
                eta_col = headers.index(h) + 1
                break
        if eta_col is None:
            continue
        for r in range(header_row + 1, ws.max_row + 1):
            wp = ws.cell(r, 1).value
            if not (isinstance(wp, str) and wp.startswith("WP")):
                continue
            eta = ws.cell(r, eta_col).value
            if eta is None or eta == "":
                continue
            if isinstance(eta, str) and not eta_pattern.match(eta.strip()):
                findings.append(Finding(
                    "error",
                    f"{tab}:{_col_letter(eta_col)}{r}",
                    f"ETA {eta!r} does not match project format "
                    f"'<Day> <H:MM> <AM/PM>' (e.g., 'Tue 3:00 PM'). "
                    f"System prompt: 'Times in 12-hour clock format with "
                    f"day prefix (e.g., \"Mon 1:50 PM\")'.",
                ))

    return findings


def _col_letter(col_idx_1based: int) -> str:
    """Convert 1-based column index to Excel letter ('A', 'B', ..., 'AA')."""
    from openpyxl.utils import get_column_letter
    return get_column_letter(col_idx_1based)


# ==============================================================
# Convenience runner with formatted console output
# ==============================================================

def print_verification_report(output_path: str, passage: dict, forecast: dict, buoys: dict) -> int:
    """Print the verification report to stdout. Returns 0 if clean, N if findings."""
    print("\n" + "=" * 70)
    print("END-OF-RUN VERIFICATION")
    print("=" * 70)
    findings = verify_workbook(output_path, passage, forecast, buoys)
    if not findings:
        print("  ✓ All checks passed. Workbook is internally consistent.")
        return 0
    errors = sum(1 for f in findings if f.severity == "error")
    warns = sum(1 for f in findings if f.severity == "warn")
    infos = sum(1 for f in findings if f.severity == "info")
    print(f"  Findings: {errors} error(s), {warns} warning(s), {infos} info")
    for f in findings:
        print(f.format())
    return errors
